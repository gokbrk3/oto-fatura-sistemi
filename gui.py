
import tkinter as tk
from tkinter import ttk, messagebox
import os
import json
import threading
import traceback
import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException

# Selenium fonksiyonlarını import et
from selenium_taslak_oluşturuyor import (
    login_portal,
    create_invoice_simple,
    check_customer_and_edit,
    upload_products_from_excel,
    add_invoice_note,
    save_and_close_invoice
)

# Global değişkenler
driver_global = None
fatura_queue = []
is_processing = False
headless_var = None
log_text = None
queue_table = None
fatura_indirme_aktif = False  # Fatura indirme durumu kontrolü
fatura_indirme_kuyrugu = []  # Fatura indirme kuyruğu
secilen_musteri_vknleri = set()  # Seçilen müşterilerin VKN'leri







# ================== START HELPERS ==================

def vkn_ile_musteri_ismi_bul(vkn):
    """
    VKN'ye göre musteriler.json'dan müşteri unvanını bulur.
    Bulunamazsa None döner.
    """
    try:
        with open("musteriler.json", "r", encoding="utf-8") as f:
            musteri_data = json.load(f)
        
        for musteri in musteri_data:
            if len(musteri) >= 4 and musteri[0] == vkn:
                # Unvan 4. sütunda (index 3)
                unvan = musteri[3].strip()
                if unvan:  # Unvan boş değilse
                    return unvan
        
        return None  # VKN bulunamadı veya unvan boş
    except Exception as e:
        log_yaz(f"⚠️ Müşteri verisi okunamadı: {e}")
        return None

class AutocompleteCombobox(tk.Frame):
    def __init__(self, master, values=None, width=20, next_widget=None,
                 linked_fields=None, kart_table=None, **kwargs):
        super().__init__(master, **kwargs)
        self.values = values if values else []
        self.next_widget = next_widget
        self.linked_fields = linked_fields
        self.kart_table = kart_table

        self.var = tk.StringVar()

        # Üst satır: Entry + ▼ butonu
        top_frame = tk.Frame(self)
        top_frame.pack(fill="x")

        self.entry = tk.Entry(top_frame, textvariable=self.var, width=width)
        self.entry.pack(side="left", fill="x", expand=True)

        self.button = tk.Button(top_frame, text="▼", width=2, command=self.show_all)
        self.button.pack(side="right")

        # Aşağı açılan liste
        self.listbox = tk.Listbox(self, height=5)
        self.listbox.bind("<<ListboxSelect>>", self.on_select)
        self.listbox.pack_forget()

        # Entry eventleri
        self.entry.bind("<KeyRelease>", self.on_keyrelease)
        self.entry.bind("<Down>", self.focus_listbox)
        self.entry.bind("<Return>", self.confirm_selection)
        self.entry.bind("<Tab>", self.confirm_selection)

        # Listbox eventleri
        self.listbox.bind("<Return>", self.confirm_selection)
        self.listbox.bind("<Double-Button-1>", self.on_select)
        self.listbox.bind("<Tab>", self.confirm_selection)
        self.listbox.bind("<Down>", self.move_down)
        self.listbox.bind("<Up>", self.move_up)

    def show_all(self):
        self.listbox.delete(0, tk.END)
        for v in self.values:
            self.listbox.insert(tk.END, v)
        self.listbox.pack(fill="x")

    def get_values(self):
        return self.values
    def set_values(self, new_values):
        self.values = new_values
    @property
    def master_values(self):
        return self.values
    @master_values.setter
    def master_values(self, new_values):
        self.values = new_values

    def on_keyrelease(self, event):
        text = self.var.get().lower()
        if text == "":
            self.listbox.pack_forget()
            return
        matches = [v for v in self.values if text in v.lower()]
        self.listbox.delete(0, tk.END)
        if matches:
            for m in matches:
                self.listbox.insert(tk.END, m)
            self.listbox.pack(fill="x")
        else:
            self.listbox.pack_forget()

    def fill_linked_fields(self, value):
        if not (self.kart_table and self.linked_fields and value):
            return
        for child in self.kart_table.get_children():
            tur, ad, b, f, k = self.kart_table.item(child, "values")
            if value == f"{tur} ({ad})":
                self.linked_fields["birim"].delete(0, "end")
                self.linked_fields["birim"].insert(0, b)
                self.linked_fields["fiyat"].delete(0, "end")
                self.linked_fields["fiyat"].insert(0, f)
                self.linked_fields["kdv"].delete(0, "end")
                self.linked_fields["kdv"].insert(0, k)
                break

    def on_select(self, event):
        if not self.listbox.curselection():
            return
        index = self.listbox.curselection()[0]
        value = self.listbox.get(index)
        self.var.set(value)
        self.listbox.pack_forget()
        self.fill_linked_fields(value)
        if self.next_widget:
            self.next_widget.focus_set()

    def confirm_selection(self, event):
        if self.listbox.curselection():
            index = self.listbox.curselection()[0]
            value = self.listbox.get(index)
        else:
            value = self.var.get()
        self.var.set(value)
        self.listbox.pack_forget()
        self.fill_linked_fields(value)
        if self.next_widget:
            self.next_widget.focus_set()
        return "break"

    def focus_listbox(self, event):
        if self.listbox.size() > 0:
            self.listbox.focus_set()
            self.listbox.selection_clear(0, tk.END)
            self.listbox.selection_set(0)
            self.listbox.activate(0)
        return "break"

    def move_down(self, event):
        idx = self.listbox.curselection()[0] if self.listbox.curselection() else -1
        if idx < self.listbox.size() - 1:
            self.listbox.selection_clear(0, tk.END)
            self.listbox.selection_set(idx + 1)
            self.listbox.activate(idx + 1)
        return "break"

    def move_up(self, event):
        idx = self.listbox.curselection()[0] if self.listbox.curselection() else 0
        if idx > 0:
            self.listbox.selection_clear(0, tk.END)
            self.listbox.selection_set(idx - 1)
            self.listbox.activate(idx - 1)
        return "break"

    def get(self):
        return self.var.get()

    def set(self, value):
        self.var.set(value)
        self.listbox.pack_forget()
        if value:
            self.fill_linked_fields(value)

    def update_values(self, new_values):
        self.values = new_values
# ================== END HELPERS ==================








# ================== START CONFIG ==================
import json
import os

AYARLAR_DOSYA = "ayarlar.json"
KARTLAR_DOSYA = "urun_kartlari.json"
MUSTERI_DOSYA = "musteriler.json"

# Varsayılan ayarlar
default_settings = {
    "pencere_boyut": "1000x750",
    "son_sekme": "Fatura"
}

# --- Ayarlar ---
def load_settings():
    if os.path.exists(AYARLAR_DOSYA):
        try:
            with open(AYARLAR_DOSYA, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return default_settings.copy()
    return default_settings.copy()

def save_settings(settings):
    with open(AYARLAR_DOSYA, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=4)
# --- Ayarlar Sonu ---

# --- Ürün Kartları ---
def save_kartlar(kart_table):
    """Ürün kartlarını kaydet - TEHLİKELİ: Sadece görünenleri kaydediyor!"""
    # UYARI: Bu fonksiyon tehlikeli! Arama yapıldığında diğer ürünler kaybolabilir!
    data = []
    for child in kart_table.get_children():
        values = kart_table.item(child, "values")
        data.append(values)
    with open(KARTLAR_DOSYA, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def load_kartlar(kart_table):
    if not os.path.exists(KARTLAR_DOSYA):
        return
    try:
        with open(KARTLAR_DOSYA, "r", encoding="utf-8") as f:
            data = json.load(f)
        for values in data:
            kart_table.insert("", "end", values=values)
        
        # Ürün listesini güncelle
        urun_listesi = []
        for child in kart_table.get_children():
            tur, ad, b, f, k = kart_table.item(child, "values")
            urun_listesi.append(f"{tur} ({ad})")
        
        # Combobox'ları güncelle (urun_combo henüz tanımlanmamış olabilir)
        try:
            if 'urun_combo' in globals():
                urun_combo['values'] = urun_listesi
        except:
            pass
    except:
        pass

def safe_add_urun_kart(yeni_urun):
    """Güvenli ürün kartı ekleme - mevcut ürünleri korur"""
    try:
        # Mevcut ürünleri oku
        data = []
        if os.path.exists(KARTLAR_DOSYA):
            with open(KARTLAR_DOSYA, "r", encoding="utf-8") as f:
                data = json.load(f)
        
        # Yeni ürünü ekle
        data.append(yeni_urun)
        
        # Dosyayı kaydet
        with open(KARTLAR_DOSYA, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        return True
    except Exception as e:
        print(f"Ürün ekleme hatası: {e}")
        return False

def safe_update_urun_kart(eski_urun, yeni_urun):
    """Güvenli ürün kartı güncelleme - sadece belirtilen ürünü günceller"""
    try:
        # Mevcut ürünleri oku
        if not os.path.exists(KARTLAR_DOSYA):
            return False
        
        with open(KARTLAR_DOSYA, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        # İlgili ürünü bul ve güncelle
        updated = False
        for i, urun in enumerate(data):
            # Ürün adı ve türüyle eşleştir
            if len(urun) >= 2 and urun[0] == eski_urun[0] and urun[1] == eski_urun[1]:
                data[i] = yeni_urun
                updated = True
                break
        
        if updated:
            # Dosyayı kaydet
            with open(KARTLAR_DOSYA, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
            return True
        return False
    except Exception as e:
        print(f"Ürün güncelleme hatası: {e}")
        return False

def safe_delete_urun_kart(silinecek_urun):
    """Güvenli ürün kartı silme - sadece belirtilen ürünü siler"""
    try:
        # Mevcut ürünleri oku
        if not os.path.exists(KARTLAR_DOSYA):
            return False
        
        with open(KARTLAR_DOSYA, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        # İlgili ürünü bul ve sil
        new_data = []
        deleted = False
        for urun in data:
            # Ürün adı ve türüyle eşleştir
            if len(urun) >= 2 and urun[0] == silinecek_urun[0] and urun[1] == silinecek_urun[1]:
                deleted = True
                continue  # Bu ürünü atlayarak sil
            new_data.append(urun)
        
        if deleted:
            # Dosyayı kaydet
            with open(KARTLAR_DOSYA, "w", encoding="utf-8") as f:
                json.dump(new_data, f, ensure_ascii=False, indent=4)
            return True
        return False
    except Exception as e:
        print(f"Ürün silme hatası: {e}")
        return False

# --- Ürün Kartları Sonu ---

# --- Müşteriler ---
def save_musteriler(musteri_table):
    """Müşteri tablosunu kaydet - SADECE seçili müşteriyi güncelle, diğerlerini koru"""
    # UYARI: Bu fonksiyon tehlikeli! Sadece tablodaki görünen müşterileri kaydediyor
    # Arama yapıldığında diğer müşteriler kaybolabilir!
    # Daha güvenli bir güncelleme sistemi gerekli
    data = [musteri_table.item(c, "values") for c in musteri_table.get_children()]
    with open(MUSTERI_DOSYA, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def safe_update_musteri(vkn, new_values):
    """Güvenli müşteri güncelleme - sadece belirtilen VKN'li müşteriyi günceller"""
    try:
        # Tüm müşterileri oku
        if not os.path.exists(MUSTERI_DOSYA):
            print(f"❌ {MUSTERI_DOSYA} dosyası bulunamadı")
            return False
        
        with open(MUSTERI_DOSYA, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        print(f"🔍 Aranan VKN: '{vkn}', Toplam müşteri sayısı: {len(data)}")
        
        # İlgili müşteriyi bul ve güncelle
        updated = False
        for i, musteri in enumerate(data):
            if len(musteri) > 0:
                # VKN'leri normalize et (başındaki sıfırları temizle)
                musteri_vkn = str(musteri[0]).strip().lstrip('0') or '0'
                aranan_vkn = str(vkn).strip().lstrip('0') or '0'
                
                if musteri_vkn == aranan_vkn:
                    print(f"✅ Müşteri bulundu! İndex: {i}, VKN: '{musteri[0]}' -> '{aranan_vkn}'")
                    data[i] = new_values
                    updated = True
                    break
        
        if updated:
            # Dosyayı kaydet
            with open(MUSTERI_DOSYA, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
            print("✅ Müşteri başarıyla güncellendi")
            return True
        else:
            print(f"❌ VKN '{vkn}' ile müşteri bulunamadı")
            return False
    except Exception as e:
        print(f"❌ Müşteri güncelleme hatası: {e}")
        return False

def safe_delete_musteri(vkn):
    """Güvenli müşteri silme - sadece belirtilen VKN'li müşteriyi siler"""
    try:
        # Tüm müşterileri oku
        if not os.path.exists(MUSTERI_DOSYA):
            return False
        
        with open(MUSTERI_DOSYA, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        # İlgili müşteriyi bul ve sil
        new_data = []
        deleted = False
        for musteri in data:
            if len(musteri) > 0:
                # VKN'leri normalize et
                musteri_vkn = str(musteri[0]).strip().lstrip('0') or '0'
                aranan_vkn = str(vkn).strip().lstrip('0') or '0'
                if musteri_vkn == aranan_vkn:
                    deleted = True
                    continue  # Bu müşteriyi atlayarak sil
            new_data.append(musteri)
        
        if deleted:
            # Dosyayı kaydet
            with open(MUSTERI_DOSYA, "w", encoding="utf-8") as f:
                json.dump(new_data, f, ensure_ascii=False, indent=4)
            return True
        return False
    except Exception as e:
        print(f"Müşteri silme hatası: {e}")
        return False

def load_musteriler(musteri_table):
    if not os.path.exists(MUSTERI_DOSYA):
        return
    try:
        with open(MUSTERI_DOSYA, "r", encoding="utf-8") as f:
            data = json.load(f)
        for values in data:
            musteri_table.insert("", "end", values=values)
    except:
        pass
# --- Müşteriler Sonu ---
# ================== END CONFIG ==================

# ================== START LOG ==================
import os

# GUI tarafındaki log_text değişkeni sonradan atanacak
log_text = None  

def log_yaz(mesaj):
    """
    Hem GUI log ekranına hem de konsola yazdırır.
    GUI tarafında log_text varsa oraya da ekler.
    """
    global log_text
    try:
        if log_text is not None:
            log_text.configure(state="normal")
            log_text.insert("end", mesaj + os.linesep)
            log_text.configure(state="disabled")
            log_text.see("end")
    except Exception:
        pass  # GUI yoksa hata verme

    print(mesaj)  # Konsola yazdır
# ================== END LOG ==================



# ================== START QUEUE VIEW ==================
def init_queue_view(frame_parent):
    global queue_table
    from tkinter import ttk

    # Kuyruk tablosu
    queue_frame = tk.LabelFrame(frame_parent, text="Fatura Kuyruğu", padx=10, pady=10, bg="#d0d0d0")
    queue_frame.pack(fill="both", expand=True, pady=10)

    columns = ("Unvan", "Vergi No", "Açıklama")
    queue_table = ttk.Treeview(queue_frame, columns=columns, show="headings", height=5)

    for col in columns:
        queue_table.heading(col, text=col)
        queue_table.column(col, width=200, anchor="center")

    queue_table.pack(fill="both", expand=True, side="left")
    scroll = ttk.Scrollbar(queue_frame, orient="vertical", command=queue_table.yview)
    queue_table.configure(yscrollcommand=scroll.set)
    scroll.pack(side="right", fill="y")
    
    # Zebra görünümü uygula
    apply_zebra_striping(queue_table)

def refresh_queue_view():
    """GUI'deki kuyruk tablosunu günceller"""
    global queue_table, fatura_queue
    if not queue_table:
        return

    queue_table.delete(*queue_table.get_children())
    for _, _, _, bilgiler in fatura_queue:
        queue_table.insert("", "end", values=(
            bilgiler.get("unvan", ""),
            bilgiler.get("vergi_no", ""),
            (bilgiler.get("aciklama", "")[:30] + "...") if bilgiler.get("aciklama") else ""
        ))
# ================== END QUEUE VIEW ==================




# ================== START ZEBRA STRIPING ==================
def apply_zebra_striping(table):
    """Tabloya zebra görünümü (çizgili satırlar) uygular"""
    table.tag_configure("even", background="white")
    table.tag_configure("odd", background="#e0e0e0")  # Daha koyu gri
    for i, item in enumerate(table.get_children()):
        if i % 2 == 0:
            table.item(item, tags=("even",))
        else:
            table.item(item, tags=("odd",))
# ================== END ZEBRA STRIPING ==================

# ================== START CONTROLLER ==================
def add_kart(tur_combo, ad_entry, birim_entry, fiyat_entry, kdv_combo,
             kart_table, urun_combo, editing_id=None):
    urun_tur = tur_combo.get().strip()
    urun_ad = ad_entry.get().strip()
    birim = birim_entry.get().strip()
    fiyat = fiyat_entry.get().strip()
    kdv = kdv_combo.get().strip()

    if not (urun_tur and urun_ad):
        return

    values = (urun_tur, urun_ad, birim, fiyat, kdv)

    if editing_id:  # Güncelleme modu
        # Eski değerleri al
        eski_values = kart_table.item(editing_id, "values")
        if safe_update_urun_kart(eski_values, values):
            kart_table.item(editing_id, values=values)
        else:
            tk.messagebox.showerror("Hata", "Ürün güncellenemedi!")
            return None
    else:  # Yeni ekleme
        # Güvenli ürün ekleme kullan
        if safe_add_urun_kart(values):
            # Tabloya da ekle
            kart_table.insert("", "end", values=values)
            full_name = f"{urun_tur} ({urun_ad})"
            current = list(urun_combo['values'])
            if full_name not in current:
                urun_combo['values'] = current + [full_name]
            
            # Zebra görünümünü yenile
            apply_zebra_striping(kart_table)
        else:
            tk.messagebox.showerror("Hata", "Ürün eklenemedi!")
            return None

    # Alanları sıfırla
    tur_combo.set("")
    ad_entry.delete(0, "end")
    birim_entry.delete(0, "end"); birim_entry.insert(0, "ADET")
    fiyat_entry.delete(0, "end")
    kdv_combo.set("20")

    return None


def add_urun(urun_combo, miktar_entry, birim_entry, fiyat_entry, kdv_entry,
             iskonto_entry, aciklama_entry, urun_table, kart_table, editing_id=None):
    urun = urun_combo.get().strip()
    miktar = miktar_entry.get().strip()
    birim = birim_entry.get().strip()
    fiyat = fiyat_entry.get().strip()
    kdv = kdv_entry.get().strip()
    iskonto = iskonto_entry.get().strip()
    aciklama = aciklama_entry.get().strip()

    if not urun:
        return

    values = (urun, miktar, birim, fiyat, kdv, iskonto, aciklama)

    if editing_id:  # Güncelleme modu
        urun_table.item(editing_id, values=values)
    else:  # Yeni ekleme
        urun_table.insert("", "end", values=values)
    
    # Zebra görünümünü yenile
    apply_zebra_striping(urun_table)

    # Alanları sıfırla
    urun_combo.set("")
    miktar_entry.delete(0, "end"); miktar_entry.insert(0, "1")
    birim_entry.delete(0, "end")
    fiyat_entry.delete(0, "end")
    kdv_entry.delete(0, "end"); kdv_entry.insert(0, "20")
    iskonto_entry.delete(0, "end"); iskonto_entry.insert(0, "0")
    aciklama_entry.delete(0, "end")

    return None
# ================== END CONTROLLER ==================

# ================== START EXCEL ==================
import os
from openpyxl import load_workbook

def _to_float_safe(val):
    if val is None or str(val).strip() == "":
        return 0.0
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except:
        return 0.0

def create_temp_excel_from_table(
    urun_table,
    template_file="zirve_excel_şablon.xlsx",
    output_file="test_fatura_zirve.xlsx"
):
    """
    GUI'deki ürün tablosunu alır, Zirve şablonuna göre geçici Excel dosyası oluşturur.
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_dir, template_file)
    output_path = os.path.join(base_dir, output_file)

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Zirve şablon dosyası bulunamadı: {template_path}")

    wb = load_workbook(template_path)
    ws = wb.active

    row_idx = 2
    for item in urun_table.get_children():
        urun, miktar, birim, fiyat, kdv, iskonto, aciklama = urun_table.item(item, "values")

        miktar_f = _to_float_safe(miktar)
        fiyat_f = _to_float_safe(fiyat)
        iskonto_yuzde = _to_float_safe(iskonto)

        # 👇 iskonto tutarını hesapla
        iskonto_tutar = round((fiyat_f * miktar_f) * (iskonto_yuzde / 100.0), 2)

        # 👇 KDV yüzde (boşsa 0)
        try:
            kdv_yuzde = int(float(kdv)) if str(kdv).strip() != "" else 0
        except:
            kdv_yuzde = 0

        # 👇 KDV tutarı hesapla (iskonto sonrası)
        kdv_tutar = round(((fiyat_f * miktar_f) - iskonto_tutar) * (kdv_yuzde / 100.0), 2)

        # 👇 Birim
        birim_yazi = "C62" if miktar_f > 0 else (birim or "")

        # Excel sütunlarına sırayla yaz
        ws[f"A{row_idx}"] = urun
        ws[f"B{row_idx}"] = miktar_f if miktar != "" else None
        ws[f"C{row_idx}"] = birim_yazi
        ws[f"D{row_idx}"] = fiyat_f
        ws[f"E{row_idx}"] = iskonto_yuzde
        ws[f"F{row_idx}"] = iskonto_tutar
        ws[f"G{row_idx}"] = kdv_yuzde
        ws[f"H{row_idx}"] = kdv_tutar
        ws[f"I{row_idx}"] = aciklama

        row_idx += 1

    wb.save(output_path)
    return output_path
# ================== END EXCEL ==================







# ================== START GUI ==================
def gui_main():
    global urun_table, kart_table
    global musteri_vkn, musteri_unvan, musteri_adi, musteri_soyadi
    global musteri_vd_sehir, musteri_vd, musteri_adres, musteri_adres_sehir, musteri_ilce
    global fatura_aciklama
    global zirve_sirket_combo, zirve_user, zirve_pass
    global log_text, headless_var
    global fatura_kes_sube_combo, fatura_kes_personel_entry, fatura_kes_islem_turu_combo
    global efatura_table, earsiv_table, secilen_musteri_vknleri
    
    root = tk.Tk()
    root.title("Oto Fatura Programı")
    root.configure(bg="#d0d0d0")  # Ana pencere arka plan rengi - daha koyu gri



    ayarlar = load_settings()
    root.geometry(ayarlar.get("pencere_boyut", "1000x750"))

    # --- Sağ Tık Silme Menüsü ---
    def attach_context_delete(table: ttk.Treeview):
        menu = tk.Menu(root, tearoff=0)
        def delete_selected():
            # Hangi tabloda olduğumuzu anla
            table_name = None
            for name, obj in globals().items():
                if obj is table:
                    table_name = name
                    break
            
            # Seçili öğeleri güvenli şekilde sil
            deleted_count = 0
            for item in table.selection():
                values = table.item(item, "values")
                if values:
                    # Tablo türüne göre güvenli silme fonksiyonu kullan
                    success = False
                    if table_name == "kart_table":
                        success = safe_delete_urun_kart(values)
                    elif table_name == "musteri_table":
                        success = safe_delete_musteri(values[0])  # VKN ile sil
                    else:
                        # Bilinmeyen tablo, eski yöntemi kullan
                        table.delete(item)
                        success = True
                    
                    if success:
                        table.delete(item)
                        deleted_count += 1
            
            if deleted_count > 0:
                tk.messagebox.showinfo("Başarılı", f"{deleted_count} öğe silindi")
            
        menu.add_command(label="Sil", command=delete_selected)
        def on_right_click(event):
            iid = table.identify_row(event.y)
            if iid:
                if iid not in table.selection():
                    table.selection_set(iid)
                menu.post(event.x_root, event.y_root)
        table.bind("<Button-3>", on_right_click)
    # --- Sağ Tık Silme Menüsü Sonu ---

    # --- Zebra Görünümü ---
    # Global apply_zebra_striping fonksiyonu kullanılacak
    # --- Zebra Görünümü Sonu ---

    # --- Notebook (Sekmeler) ---
    # Notebook'u gri Frame ile sarmalayalım
    notebook_frame = tk.Frame(root, bg="#d0d0d0")
    notebook_frame.pack(fill="both", expand=True)
    
    # Notebook stilini önce ayarla
    style = ttk.Style()
    style.theme_use('clam')  # Tema değiştir
    style.configure("TNotebook", background="#d0d0d0", borderwidth=0)
    style.configure("TNotebook.Tab", background="#d0d0d0", foreground="black", borderwidth=0)
    style.map("TNotebook.Tab", background=[("selected", "white"), ("active", "#e0e0e0")])
    
    # Zebra görünümü için Treeview stilleri
    style.configure("Treeview", background="white", foreground="black", fieldbackground="white")
    style.configure("Treeview.Heading", background="#d0d0d0", foreground="black")
    style.map("Treeview", background=[("selected", "#0078d4")])
    
    notebook = ttk.Notebook(notebook_frame, style="TNotebook")
    notebook.pack(fill="both", expand=True, padx=2, pady=2)
    
    # Notebook'u zorla gri yap
    try:
        notebook.configure(style="TNotebook")
        notebook.tk.call("ttk::style", "configure", "TNotebook", "-background", "#d0d0d0")
        notebook.tk.call("ttk::style", "configure", "TNotebook.Tab", "-background", "#d0d0d0")
    except:
        pass
    
    # Entry ve Text widget'ları beyaz kalacak, sadece label'lar gri olacak
    # Tkinter Entry ve Text widget'ları beyaz kalacak (varsayılan)
    # Label'lar için arka plan rengi ayarla
    root.option_add("*Label*background", "#d0d0d0")
    # --- Notebook Sonu ---

    # ===================================================
    # ==== SEKME 1: FATURA ====
    # ===================================================
    frame_fatura = tk.Frame(notebook, padx=10, pady=10, bg="#d0d0d0")
    notebook.add(frame_fatura, text="Fatura Taslak Oluştur")

    # --- Müşteri Bilgileri ---
    frame_musteri = tk.LabelFrame(frame_fatura, text="Müşteri Bilgileri", padx=10, pady=10, bg="#d0d0d0")
    frame_musteri.pack(fill="x", pady=10)

    # Satır 0: VKN/TCKN, Unvan, Adı, Soyadı, Adres (2 satır kaplar), Şubeler
    tk.Label(frame_musteri, text="VKN / TCKN:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
    musteri_vkn = tk.Entry(frame_musteri, width=18)
    musteri_vkn.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="Unvan:").grid(row=0, column=2, sticky="e", padx=5, pady=5)
    musteri_unvan = tk.Entry(frame_musteri, width=20)
    musteri_unvan.grid(row=0, column=3, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="Adı:").grid(row=0, column=4, sticky="e", padx=5, pady=5)
    musteri_adi = tk.Entry(frame_musteri, width=15)
    musteri_adi.grid(row=0, column=5, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="Soyadı:").grid(row=0, column=6, sticky="e", padx=5, pady=5)
    musteri_soyadi = tk.Entry(frame_musteri, width=15)
    musteri_soyadi.grid(row=0, column=7, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="Adres:").grid(row=0, column=8, sticky="ne", padx=5, pady=5)
    musteri_adres = tk.Text(frame_musteri, width=35, height=4)
    musteri_adres.grid(row=0, column=9, rowspan=2, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="Şubeler:").grid(row=0, column=10, sticky="e", padx=5, pady=5)
    musteri_subeler = tk.Entry(frame_musteri, width=25)
    musteri_subeler.grid(row=0, column=11, padx=5, pady=5, sticky="w")

    # Satır 1: Vergi D. Şehir, Vergi Dairesi, Adres Şehir, İlçe
    tk.Label(frame_musteri, text="Vergi D. Şehir:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
    musteri_vd_sehir = tk.Entry(frame_musteri, width=18)
    musteri_vd_sehir.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="Vergi Dairesi:").grid(row=1, column=2, sticky="e", padx=5, pady=5)
    musteri_vd = tk.Entry(frame_musteri, width=20)
    musteri_vd.grid(row=1, column=3, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="Adres Şehir:").grid(row=1, column=4, sticky="e", padx=5, pady=5)
    musteri_adres_sehir = tk.Entry(frame_musteri, width=15)
    musteri_adres_sehir.grid(row=1, column=5, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="İlçe:").grid(row=1, column=6, sticky="e", padx=5, pady=5)
    musteri_ilce = tk.Entry(frame_musteri, width=15)
    musteri_ilce.grid(row=1, column=7, padx=5, pady=5, sticky="w")

    # --- Fonksiyonlar ---
    def kaydet_musteri():
        values = (
            musteri_vkn.get(),
            musteri_adi.get(),
            musteri_soyadi.get(),
            musteri_unvan.get(),
            musteri_vd_sehir.get(),
            musteri_vd.get(),
            musteri_adres_sehir.get(),
            musteri_ilce.get(),
            musteri_subeler.get(),
            musteri_adres.get("1.0", "end").strip()
        )

        if not values[0]:
            return

        # VKN kontrolü
        for child in musteri_table.get_children():
            mevcut = musteri_table.item(child, "values")
            if mevcut[0] == values[0]:
                # Aynı VKN bulundu → bilgileri getir
                messagebox.showinfo("Bilgi", "Bu VKN/TCKN zaten kayıtlı, bilgileri dolduruldu.")
                musteri_vkn.delete(0, "end"); musteri_vkn.insert(0, mevcut[0])
                musteri_adi.delete(0, "end"); musteri_adi.insert(0, mevcut[1])
                musteri_soyadi.delete(0, "end"); musteri_soyadi.insert(0, mevcut[2])
                musteri_unvan.delete(0, "end"); musteri_unvan.insert(0, mevcut[3])
                musteri_vd_sehir.delete(0, "end"); musteri_vd_sehir.insert(0, mevcut[4])
                musteri_vd.delete(0, "end"); musteri_vd.insert(0, mevcut[5])
                musteri_adres_sehir.delete(0, "end"); musteri_adres_sehir.insert(0, mevcut[6])
                musteri_ilce.delete(0, "end"); musteri_ilce.insert(0, mevcut[7])
                musteri_subeler.delete(0, "end"); musteri_subeler.insert(0, mevcut[8])
                musteri_adres.delete("1.0", "end"); musteri_adres.insert("1.0", mevcut[9])

                # Şube combobox doldur
                subeler_list = [s.strip() for s in mevcut[8].split(",") if s.strip()]
                # sube_combo kaldırıldı
                # sube_combo kaldırıldı
                return

        # Yeni müşteri ekleniyor
        musteri_table.insert("", "end", values=values)
        messagebox.showinfo("Bilgi", "Yeni müşteri kaydedildi.")

        try:
            save_musteriler(musteri_table)
        except Exception:
            pass

        # Şube combobox doldur
        subeler_list = [s.strip() for s in values[8].split(",") if s.strip()]
        # sube_combo kaldırıldı

    def musteri_cagir():
        win = tk.Toplevel(root)
        win.title("Müşteri Seç")
        win.geometry("900x400")

        search_var = tk.StringVar()
        tk.Label(win, text="Ara:").pack(anchor="w", padx=5, pady=5)
        search_entry = tk.Entry(win, textvariable=search_var, width=50)
        search_entry.pack(fill="x", padx=5, pady=5)

        columns = (
            "VKN/TCKN", "Adı", "Soyadı", "Unvan", "Vergi D. Şehir",
            "Vergi Dairesi", "Adres Şehir", "İlçe", "Şubeler", "Adres"
        )
        table = ttk.Treeview(win, columns=columns, show="headings", height=12)
        for col in columns:
            table.heading(col, text=col)
            table.column(col, width=120, anchor="center")
        table.pack(fill="both", expand=True, side="left", padx=5, pady=5)
        # Scrollbar
        scroll = ttk.Scrollbar(win, orient="vertical", command=table.yview)
        table.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")

        for child in musteri_table.get_children():
            table.insert("", "end", values=musteri_table.item(child, "values"))

        def do_search(*args):
            query = search_var.get().lower()
            table.delete(*table.get_children())
            for child in musteri_table.get_children():
                values = musteri_table.item(child, "values")
                if any(query in str(v).lower() for v in values):
                    table.insert("", "end", values=values)
        search_var.trace("w", do_search)

        def select_customer():
            selected = table.selection()
            if not selected:
                messagebox.showinfo("Uyarı", "Lütfen bir müşteri seçin!")
                return
            values = table.item(selected[0], "values")
            if not values:
                return

            musteri_vkn.delete(0, "end"); musteri_vkn.insert(0, values[0])
            musteri_adi.delete(0, "end"); musteri_adi.insert(0, values[1])
            musteri_soyadi.delete(0, "end"); musteri_soyadi.insert(0, values[2])
            musteri_unvan.delete(0, "end"); musteri_unvan.insert(0, values[3])
            musteri_vd_sehir.delete(0, "end"); musteri_vd_sehir.insert(0, values[4])
            musteri_vd.delete(0, "end"); musteri_vd.insert(0, values[5])
            musteri_adres_sehir.delete(0, "end"); musteri_adres_sehir.insert(0, values[6])
            musteri_ilce.delete(0, "end"); musteri_ilce.insert(0, values[7])
            musteri_subeler.delete(0, "end"); musteri_subeler.insert(0, values[8])
            musteri_adres.delete("1.0", "end"); musteri_adres.insert("1.0", values[9])

            # Şube combobox doldur
            subeler_list = [s.strip() for s in values[8].split(",") if s.strip()]
            # sube_combo kaldırıldı

            win.destroy()

        tk.Button(win, text="Seç", command=select_customer).pack(pady=10)

    def temizle_musteri():
        musteri_vkn.delete(0, "end")
        musteri_unvan.delete(0, "end")
        musteri_adi.delete(0, "end")
        musteri_soyadi.delete(0, "end")
        musteri_vd_sehir.delete(0, "end")
        musteri_vd.delete(0, "end")
        musteri_adres_sehir.delete(0, "end")
        musteri_ilce.delete(0, "end")
        musteri_subeler.delete(0, "end")
        musteri_adres.delete("1.0", "end")

    # --- Butonlar (yan yana, sağ tarafa hizalı) ---
    btn_frame = tk.Frame(frame_musteri, bg="#d0d0d0")
    btn_frame.grid(row=1, column=8, columnspan=4, padx=5, pady=5, sticky="e")

    tk.Button(btn_frame, text="Kaydet", command=kaydet_musteri).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Çağır", command=musteri_cagir).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Temizle", command=temizle_musteri).pack(side="left", padx=5)
    # --- Müşteri Bilgileri Sonu ---



    # --- Ürün Tablosu ---
    frame_urun = tk.LabelFrame(frame_fatura, text="Ürünler", padx=10, pady=10, bg="#d0d0d0")
    frame_urun.pack(fill="both", expand=True, padx=10, pady=5)

    columns = ("Ürün Adı", "Miktar", "Birim", "Birim Fiyat", "KDV %", "İskonto %", "Açıklama")
    urun_table = ttk.Treeview(frame_urun, columns=columns, show="headings", height=8)
    for col in columns:
        urun_table.heading(col, text=col)
        urun_table.column(col, width=100, anchor="center")
    urun_table.pack(fill="both", expand=True, side="left")

    scroll = ttk.Scrollbar(frame_urun, orient="vertical", command=urun_table.yview)
    urun_table.configure(yscrollcommand=scroll.set)
    scroll.pack(side="right", fill="y")
    attach_context_delete(urun_table)

    # Zebra görünümü uygula
    apply_zebra_striping(urun_table)

    # Inline fiyat düzenleme fonksiyonu
    def edit_price(event):
        item = urun_table.selection()[0] if urun_table.selection() else None
        if not item:
            return
        
        # Mevcut fiyatı al
        values = list(urun_table.item(item, "values"))
        current_price = values[3]  # Birim Fiyat
        
        # Hücrenin konumunu bul
        bbox = urun_table.bbox(item, "#4")  # "Birim Fiyat" sütunu
        if not bbox:
            return
        
        # Entry widget'ı oluştur
        edit_entry = tk.Entry(urun_table, font=("Arial", 9))
        edit_entry.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
        edit_entry.insert(0, current_price)
        edit_entry.select_range(0, tk.END)
        edit_entry.focus()
        
        def save_price():
            try:
                new_price = float(edit_entry.get())
                values[3] = str(new_price)
                urun_table.item(item, values=values)
                edit_entry.destroy()
            except ValueError:
                tk.messagebox.showerror("Hata", "Geçerli bir sayı giriniz!")
                edit_entry.destroy()
        
        def cancel_edit():
            edit_entry.destroy()
        
        # Event'leri bağla
        edit_entry.bind("<Return>", lambda e: save_price())
        edit_entry.bind("<Escape>", lambda e: cancel_edit())
        edit_entry.bind("<FocusOut>", lambda e: save_price())  # Başka yere tıklayınca kaydet

    # Çift tıklama event'ini bağla
    urun_table.bind("<Double-1>", edit_price)

    # --- Ürün Tablosu Sonu ---

    # --- Ürün Ekleme Alanı ---
    frame_add = tk.LabelFrame(frame_fatura, text="Ürün Ekleme Alanı", padx=10, pady=10, bg="#d0d0d0")
    frame_add.pack(fill="x", pady=10)

    tk.Label(frame_add, text="Ürün:").grid(row=0, column=0, sticky="w")
    urun_combo = ttk.Combobox(frame_add, values=[], width=25)
    
    # Linked fields için özel değişkenler
    urun_combo.linked_fields = {
            "birim": None,
            "fiyat": None,
            "kdv": None
    }
    urun_combo.kart_table = None
    urun_combo.grid(row=1, column=0, padx=5)

    tk.Label(frame_add, text="Miktar:").grid(row=0, column=1, sticky="w")
    miktar_entry = tk.Entry(frame_add, width=5); miktar_entry.insert(0, "1")
    miktar_entry.grid(row=1, column=1, padx=5)

    tk.Label(frame_add, text="Birim:").grid(row=0, column=2, sticky="w")
    birim_entry = tk.Entry(frame_add, width=8)
    birim_entry.grid(row=1, column=2, padx=5)

    tk.Label(frame_add, text="Fiyat:").grid(row=0, column=3, sticky="w")
    fiyat_entry = tk.Entry(frame_add, width=10)
    fiyat_entry.grid(row=1, column=3, padx=5)

    tk.Label(frame_add, text="KDV %:").grid(row=0, column=4, sticky="w")
    kdv_entry = tk.Entry(frame_add, width=5); kdv_entry.insert(0, "20")
    kdv_entry.grid(row=1, column=4, padx=5)

    tk.Label(frame_add, text="İskonto %:").grid(row=0, column=5, sticky="w")
    iskonto_entry = tk.Entry(frame_add, width=5); iskonto_entry.insert(0, "0")
    iskonto_entry.grid(row=1, column=5, padx=5)

    tk.Label(frame_add, text="Açıklama:").grid(row=0, column=6, sticky="w")
    aciklama_entry = tk.Entry(frame_add, width=20)
    aciklama_entry.grid(row=1, column=6, padx=5)

    # --- bağlantıları tamamla ---
    urun_combo.linked_fields = {
        "birim": birim_entry,
        "fiyat": fiyat_entry,
        "kdv": kdv_entry
    }
    
    # Normal Combobox için linked fields fonksiyonu
    def on_urun_selection(event):
        selected_value = urun_combo.get()
        if not (urun_combo.kart_table and urun_combo.linked_fields and selected_value):
            return
        for child in urun_combo.kart_table.get_children():
            tur, ad, b, f, k = urun_combo.kart_table.item(child, "values")
            if selected_value == f"{tur} ({ad})":
                urun_combo.linked_fields["birim"].delete(0, "end")
                urun_combo.linked_fields["birim"].insert(0, b)
                urun_combo.linked_fields["fiyat"].delete(0, "end")
                urun_combo.linked_fields["fiyat"].insert(0, f)
                urun_combo.linked_fields["kdv"].delete(0, "end")
                urun_combo.linked_fields["kdv"].insert(0, k)
                break
    
    urun_combo.bind("<<ComboboxSelected>>", on_urun_selection)

    def on_add_button():
        add_urun(
            urun_combo, miktar_entry, birim_entry, fiyat_entry, kdv_entry,
            iskonto_entry, aciklama_entry, urun_table, kart_table
        )

    tk.Button(frame_add, text="Ürün Ekle", command=on_add_button).grid(row=1, column=7, padx=5)
    tk.Button(frame_add, text="Toplu Ürün Girişi", command=lambda: open_bulk_add_window()).grid(row=1, column=8, padx=5)
    # --- Ürün Ekleme Alanı Sonu ---



    # --- Toplu Ürün Girişi ---
    def open_bulk_add_window():
        bulk_win = tk.Toplevel(root)
        bulk_win.title("Toplu Ürün Girişi")
        bulk_win.geometry("600x400")

        search_var = tk.StringVar()
        selected_items = set()

        product_listbox = tk.Listbox(bulk_win, selectmode=tk.MULTIPLE)
        product_listbox.pack(fill="both", expand=True, padx=5, pady=5)

        # Ürünleri kartlardan al
        all_products = []
        kart_map = {}
        for child in kart_table.get_children():
            tur, ad, b, f, k = kart_table.item(child, "values")
            full_name = f"{tur} ({ad})"
            all_products.append(full_name)
            kart_map[full_name] = (tur, ad, b, f, k)

        # Listeyi doldur (seçimleri koruyarak)
        def refresh_list():
            product_listbox.delete(0, tk.END)
            for p in all_products:
                if search_var.get().lower() in p.lower():
                    product_listbox.insert(tk.END, p)
                    if p in selected_items:
                        product_listbox.selection_set(tk.END)

        # Seçim değişince hafızaya al
        def on_select(event=None):
            selected = [product_listbox.get(i) for i in product_listbox.curselection()]
            for p in all_products:
                if search_var.get().lower() in p.lower():
                    if p in selected:
                        selected_items.add(p)
                    elif p in selected_items:
                        selected_items.remove(p)

        product_listbox.bind("<<ListboxSelect>>", on_select)

        # Arama kutusu
        tk.Entry(bulk_win, textvariable=search_var).pack(fill="x", padx=5, pady=5)
        search_var.trace("w", lambda *args: refresh_list())

        refresh_list()

        def add_selected_products():
            if not selected_items:
                return
            for full_name in selected_items:
                tur, ad, b, f, k = kart_map[full_name]
                urun_table.insert("", "end", values=(
                    full_name, "1", b, f, k, "0", ""
                ))
            
            # Zebra görünümünü yenile
            apply_zebra_striping(urun_table)
            
            bulk_win.destroy()

        tk.Button(bulk_win, text="Faturaya İlave Et", command=add_selected_products).pack(pady=10)
    # --- Fatura Taslak Oluştur Aksiyonu ---

 

    # --- Toplu Ürün Girişi Sonu ---

    # ================== START FATURA GENEL AÇIKLAMA ==================
    # --- Fatura Genel Açıklama ---
    frame_fatura_aciklama = tk.LabelFrame(frame_fatura, text="Fatura Açıklaması", padx=10, pady=10, bg="#d0d0d0")
    frame_fatura_aciklama.pack(fill="x", pady=10)

    fatura_aciklama = tk.Text(frame_fatura_aciklama, width=100, height=3)
    fatura_aciklama.pack(fill="x", padx=5, pady=5)
    # --- Fatura Genel Açıklama Sonu ---
    # ================== END FATURA GENEL AÇIKLAMA ==================

        # ================== START FATURA TASLAK BUTONU ==================
    btn_frame_fatura = tk.Frame(frame_fatura, bg="#d0d0d0")
    btn_frame_fatura.pack(pady=5)

    btn_fatura_olustur = tk.Button(
        btn_frame_fatura,
        text="Fatura Taslak Oluştur",
        command=lambda: log_yaz("⚠️ Fatura Kes özelliği kaldırıldı"),  # Fatura Kes özelliği kaldırıldı
        bg="green",
        fg="white"
    )
    btn_fatura_olustur.pack(side="left", padx=5, pady=5)
    # ================== END FATURA TASLAK BUTONU ==================



    # ===================================================
    # ==== SEKME 2: FATURA İNDİR ====
    # ===================================================
    frame_fatura_indir = tk.Frame(notebook, padx=10, pady=10, bg="#d0d0d0")
    notebook.add(frame_fatura_indir, text="Fatura İndir")
    
    # --- Fatura İsimlendirme ---
    frame_fatura_secim = tk.LabelFrame(frame_fatura_indir, text="Fatura İsimlendirme", padx=10, pady=10, bg="#d0d0d0")
    frame_fatura_secim.pack(fill="x", pady=10)
    
    # Tek satır: Şube, Personel, İşlem Türü, Fatura İndir (büyük textboxlar ile)
    tk.Label(frame_fatura_secim, text="Şube:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
    global fatura_kes_sube_combo
    fatura_kes_sube_combo = ttk.Combobox(frame_fatura_secim, values=[], width=24)  # 12'den 24'e çıkarıldı
    fatura_kes_sube_combo.grid(row=0, column=1, padx=5, pady=5)
    
    # Şube ekleme butonu - çok küçük gri "+" butonu
    def sube_ekle():
        """Manuel girilen şubeyi seçilen müşterilerin şubeler alanına ekle"""
        yeni_sube = fatura_kes_sube_combo.get().strip()
        if not yeni_sube:
            tk.messagebox.showwarning("Uyarı", "Lütfen şube adı giriniz!")
            return
        
        # Şubeyi müşterilere ekle
        sube_musterilere_ekle(yeni_sube)
    
    btn_sube_ekle = tk.Button(frame_fatura_secim, text="+", command=sube_ekle, 
                             bg="#E8E8E8", fg="black", font=("Arial", 8, "bold"),
                             width=1, height=1, relief="solid", bd=1)
    btn_sube_ekle.grid(row=0, column=2, padx=2, pady=5)
    
    tk.Label(frame_fatura_secim, text="Personel:").grid(row=0, column=3, sticky="w", padx=(15,5), pady=5)
    fatura_kes_personel_entry = tk.Entry(frame_fatura_secim, width=24)  # 12'den 24'e çıkarıldı
    fatura_kes_personel_entry.grid(row=0, column=4, padx=5, pady=5)
    
    tk.Label(frame_fatura_secim, text="İşlem Türü:").grid(row=0, column=5, sticky="w", padx=(15,5), pady=5)
    fatura_kes_islem_turu_combo = ttk.Combobox(frame_fatura_secim, values=["SRV", "STŞ", "YEDEK PARÇA"], width=12)
    fatura_kes_islem_turu_combo.set("SRV")
    fatura_kes_islem_turu_combo.grid(row=0, column=6, padx=5, pady=5)
    
    # Fatura İndir butonu - aynı satırda, sağda (daha geniş aralık)
    btn_fatura_indir = tk.Button(frame_fatura_secim, text="Fatura İndir", command=lambda: indir_secilen_faturalar(), 
                                bg="#4CAF50", fg="white", font=("Arial", 9, "normal"), 
                                relief="raised", bd=2, padx=10, pady=5)
    btn_fatura_indir.grid(row=0, column=7, padx=15, pady=5)
    
    # --- Faturaları Oku Butonu (En altta, küçük) ---
    btn_faturalari_oku = tk.Button(frame_fatura_indir, text="Faturaları Oku", command=lambda: read_invoices_from_zirve(), 
                                  bg="#4CAF50", fg="white", font=("Arial", 9, "normal"), 
                                  relief="raised", bd=2, padx=10, pady=5)
    btn_faturalari_oku.pack(side="bottom", pady=5)
    
    # --- E-Fatura Tablosu ---
    frame_efatura = tk.LabelFrame(frame_fatura_indir, text="E-Fatura Listesi", padx=10, pady=10, bg="#d0d0d0")
    frame_efatura.pack(fill="both", expand=True, pady=5)
    
    # E-Fatura tablosu
    efatura_columns = ("Müşteri", "Vergi No", "Tutar", "Durum", "Tarih", "Fatura No")
    global efatura_table
    efatura_table = ttk.Treeview(frame_efatura, columns=efatura_columns, show="headings", height=8)
    
    for col in efatura_columns:
        efatura_table.heading(col, text=col)
        efatura_table.column(col, width=120, anchor="center")
    
    efatura_table.pack(fill="both", expand=True, side="left", pady=5)
    scroll_ef = ttk.Scrollbar(frame_efatura, orient="vertical", command=efatura_table.yview)
    efatura_table.configure(yscrollcommand=scroll_ef.set)
    scroll_ef.pack(side="right", fill="y")
    attach_context_delete(efatura_table)
    apply_zebra_striping(efatura_table)
    
    # E-Fatura tablosu seçim event'i
    efatura_table.bind("<<TreeviewSelect>>", lambda e: guncelle_subeler())
    
    # --- E-Arşiv Tablosu ---
    frame_earsiv = tk.LabelFrame(frame_fatura_indir, text="E-Arşiv Listesi", padx=10, pady=10, bg="#d0d0d0")
    frame_earsiv.pack(fill="both", expand=True, pady=5)
    
    # E-Arşiv tablosu
    earsiv_columns = ("Müşteri", "Vergi No", "Tutar", "Durum", "Tarih", "Fatura No")
    global earsiv_table
    earsiv_table = ttk.Treeview(frame_earsiv, columns=earsiv_columns, show="headings", height=8)
    
    for col in earsiv_columns:
        earsiv_table.heading(col, text=col)
        earsiv_table.column(col, width=120, anchor="center")
    
    earsiv_table.pack(fill="both", expand=True, side="left", pady=5)
    scroll_ea = ttk.Scrollbar(frame_earsiv, orient="vertical", command=earsiv_table.yview)
    earsiv_table.configure(yscrollcommand=scroll_ea.set)
    scroll_ea.pack(side="right", fill="y")
    attach_context_delete(earsiv_table)
    apply_zebra_striping(earsiv_table)
    
    # E-Arşiv tablosu seçim event'i
    earsiv_table.bind("<<TreeviewSelect>>", lambda e: guncelle_subeler())

    # ===================================================
    # ==== SEKME 3: ÜRÜN KARTLARI ====
    # ===================================================
    frame_kartlar = tk.Frame(notebook, padx=10, pady=10, bg="#d0d0d0")
    notebook.add(frame_kartlar, text="Ürün Kartları")

    # --- Ürün Kart Tablosu ---
    tk.Label(frame_kartlar, text="Ara:").pack(anchor="w", padx=5, pady=2)
    search_var_kart = tk.StringVar()
    search_entry_kart = tk.Entry(frame_kartlar, textvariable=search_var_kart, width=50)
    search_entry_kart.pack(fill="x", padx=5, pady=2)

    kart_columns = ("Ürün Türü", "Ürün Adı", "Birim", "Fiyat", "KDV %")
    # İç kapsayıcı: tablo + scrollbar yan yana
    kart_table_container = tk.Frame(frame_kartlar, bg="#d0d0d0")
    kart_table_container.pack(fill="both", expand=True, pady=10)

    kart_table = ttk.Treeview(kart_table_container, columns=kart_columns, show="headings", height=10)

    for col in kart_columns:
        kart_table.heading(col, text=col)
        kart_table.column(col, width=150, anchor="center")

    kart_table.pack(fill="both", expand=True, side="left")
    scroll_kart = ttk.Scrollbar(kart_table_container, orient="vertical", command=kart_table.yview)
    kart_table.configure(yscrollcommand=scroll_kart.set)
    scroll_kart.pack(side="right", fill="y")
    attach_context_delete(kart_table)
    load_kartlar(kart_table)
    
    # Zebra görünümü uygula
    apply_zebra_striping(kart_table)

    # 👇 Ürün Ekleme Alanı'ndaki combobox'a ürünleri yükle
    urun_listesi = []
    for child in kart_table.get_children():
        tur, ad, b, f, k = kart_table.item(child, "values")
        urun_listesi.append(f"{tur} ({ad})")
    urun_combo['values'] = urun_listesi

    # 👇 bağlantıyı burada yapıyoruz
    urun_combo.kart_table = kart_table



    # --- Ürün Kart Tablosu Arama ---
    def filter_kartlar(*args):
        query = search_var_kart.get().lower()
        kart_table.delete(*kart_table.get_children())
        try:
            with open(KARTLAR_DOSYA, "r", encoding="utf-8") as f:
                data = json.load(f)
            for values in data:
                if any(query in str(v).lower() for v in values):
                    kart_table.insert("", "end", values=values)
        except:
            pass
    search_var_kart.trace("w", filter_kartlar)
    # --- Ürün Kart Tablosu Sonu ---


    # --- Ürün Kart Ekleme Alanı ---
    frame_kart_add = tk.LabelFrame(frame_kartlar, text="Kart Ekleme Alanı", padx=10, pady=10, bg="#d0d0d0")
    frame_kart_add.pack(fill="x", pady=5)

    tk.Label(frame_kart_add, text="Ürün Türü:").grid(row=0, column=0, sticky="w")
    kart_tur_combo = ttk.Combobox(frame_kart_add, values=[
        "YEDEK PARÇA",
        "S.R/O SU ARITMA CİHAZI",
        "SEBİL",
        "R/O BETA SYSTEM - ENDÜSTRİYEL SU ARITMA CİHAZI"
    ], width=27)
    kart_tur_combo.grid(row=1, column=0, padx=5)

    tk.Label(frame_kart_add, text="Ürün Adı:").grid(row=0, column=1, sticky="w")
    kart_ad_entry = tk.Entry(frame_kart_add, width=25)
    kart_ad_entry.grid(row=1, column=1, padx=5)

    tk.Label(frame_kart_add, text="Birim:").grid(row=0, column=2, sticky="w")
    kart_birim_entry = tk.Entry(frame_kart_add, width=10); kart_birim_entry.insert(0, "ADET")
    kart_birim_entry.grid(row=1, column=2, padx=5)

    tk.Label(frame_kart_add, text="Fiyat:").grid(row=0, column=3, sticky="w")
    kart_fiyat_entry = tk.Entry(frame_kart_add, width=10)
    kart_fiyat_entry.grid(row=1, column=3, padx=5)

    tk.Label(frame_kart_add, text="KDV %:").grid(row=0, column=4, sticky="w")
    kart_kdv_combo = ttk.Combobox(frame_kart_add, values=["0", "1", "8", "10", "18", "20"], width=5)
    kart_kdv_combo.set("20")
    kart_kdv_combo.grid(row=1, column=4, padx=5)

    def on_kart_add_button():
        add_kart(
            kart_tur_combo, kart_ad_entry, kart_birim_entry,
            kart_fiyat_entry, kart_kdv_combo, kart_table, urun_combo
        )

    tk.Button(frame_kart_add, text="Kart Ekle", command=on_kart_add_button).grid(row=1, column=5, padx=10)
    
    # Düzenleme butonu
    def duzenle_kart():
        selected = kart_table.selection()
        if not selected:
            tk.messagebox.showwarning("Uyarı", "Lütfen düzenlemek için bir ürün kartı seçin!")
            return
        
        values = kart_table.item(selected[0], "values")
        if not values:
            return
        
        # Yeni düzenleme penceresi oluştur
        edit_win = tk.Toplevel(root)
        edit_win.title("Ürün Kartı Düzenle")
        edit_win.geometry("500x300")
        edit_win.grab_set()  # Modal pencere yap
        
        # Pencere içeriği
        frame_edit = tk.LabelFrame(edit_win, text="Ürün Kartı Düzenle", padx=10, pady=10)
        frame_edit.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Form alanları
        tk.Label(frame_edit, text="Ürün Türü:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        edit_tur_combo = ttk.Combobox(frame_edit, values=[
            "YEDEK PARÇA",
            "S.R/O SU ARITMA CİHAZI",
            "SEBİL",
            "R/O BETA SYSTEM - ENDÜSTRİYEL SU ARITMA CİHAZI"
        ], width=27)
        edit_tur_combo.grid(row=0, column=1, padx=5, pady=5)
        edit_tur_combo.set(values[0])
        
        tk.Label(frame_edit, text="Ürün Adı:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        edit_ad_entry = tk.Entry(frame_edit, width=30)
        edit_ad_entry.grid(row=1, column=1, padx=5, pady=5)
        edit_ad_entry.insert(0, values[1])
        
        tk.Label(frame_edit, text="Birim:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        edit_birim_entry = tk.Entry(frame_edit, width=30)
        edit_birim_entry.grid(row=2, column=1, padx=5, pady=5)
        edit_birim_entry.insert(0, values[2])
        
        tk.Label(frame_edit, text="Fiyat:").grid(row=3, column=0, sticky="w", padx=5, pady=5)
        edit_fiyat_entry = tk.Entry(frame_edit, width=30)
        edit_fiyat_entry.grid(row=3, column=1, padx=5, pady=5)
        edit_fiyat_entry.insert(0, values[3])
        
        tk.Label(frame_edit, text="KDV %:").grid(row=4, column=0, sticky="w", padx=5, pady=5)
        edit_kdv_combo = ttk.Combobox(frame_edit, values=["0", "1", "8", "10", "18", "20"], width=27)
        edit_kdv_combo.grid(row=4, column=1, padx=5, pady=5)
        edit_kdv_combo.set(values[4])
        
        # Butonlar
        btn_frame = tk.Frame(frame_edit, bg="#d0d0d0")
        btn_frame.grid(row=5, column=0, columnspan=2, pady=20)
        
        def kaydet_degisiklikler():
            # Yeni değerleri al
            yeni_tur = edit_tur_combo.get().strip()
            yeni_ad = edit_ad_entry.get().strip()
            yeni_birim = edit_birim_entry.get().strip()
            yeni_fiyat = edit_fiyat_entry.get().strip()
            yeni_kdv = edit_kdv_combo.get().strip()
            
            if not (yeni_tur and yeni_ad):
                tk.messagebox.showwarning("Uyarı", "Ürün türü ve adı zorunludur!")
                return
            
            # Eski ürün bilgilerini al
            eski_values = kart_table.item(selected[0], "values")
            yeni_values = (yeni_tur, yeni_ad, yeni_birim, yeni_fiyat, yeni_kdv)
            
            # Güvenli güncelleme kullan
            if safe_update_urun_kart(eski_values, yeni_values):
                # Tabloda da güncelle
                kart_table.item(selected[0], values=yeni_values)
            else:
                tk.messagebox.showerror("Hata", "Ürün güncellenemedi!")
                return
            
            # Ürün listesini güncelle
            urun_listesi = []
            for child in kart_table.get_children():
                tur, ad, b, f, k = kart_table.item(child, "values")
                urun_listesi.append(f"{tur} ({ad})")
            urun_combo['values'] = urun_listesi
            
            tk.messagebox.showinfo("Başarılı", "Ürün kartı başarıyla güncellendi!")
            edit_win.destroy()
        
        def iptal_et():
            edit_win.destroy()
        
        tk.Button(btn_frame, text="Kaydet", command=kaydet_degisiklikler, bg="green", fg="white").pack(side="left", padx=5)
        tk.Button(btn_frame, text="İptal", command=iptal_et, bg="red", fg="white").pack(side="left", padx=5)
    
    tk.Button(frame_kart_add, text="Düzenle", command=duzenle_kart, bg="orange", fg="white").grid(row=1, column=6, padx=10)
    # --- Ürün Kart Ekleme Alanı Sonu ---

    # ===================================================
    # --- Müşteriler Sekmesi ---
    frame_musteriler = tk.Frame(notebook, padx=10, pady=10, bg="#d0d0d0")
    notebook.add(frame_musteriler, text="Müşteriler")

    # Arama kutusu
    search_var_musteri = tk.StringVar()
    tk.Label(frame_musteriler, text="Ara:").pack(anchor="w", padx=5, pady=2)
    search_entry_musteri = tk.Entry(frame_musteriler, textvariable=search_var_musteri, width=50)
    search_entry_musteri.pack(fill="x", padx=5, pady=2)

    musteri_columns = (
        "VKN/TCKN", "Adı", "Soyadı", "Unvan", "Vergi D. Şehir",
        "Vergi Dairesi", "Adres Şehir", "İlçe", "Şubeler", "Adres"
    )
    # İç kapsayıcı: tablo + scrollbar yan yana
    musteri_table_container = tk.Frame(frame_musteriler, bg="#d0d0d0")
    musteri_table_container.pack(fill="both", expand=True, pady=10)

    musteri_table = ttk.Treeview(musteri_table_container, columns=musteri_columns, show="headings", height=12)

    for col in musteri_columns:
        musteri_table.heading(col, text=col)
        musteri_table.column(col, width=120, anchor="center")

    musteri_table.pack(fill="both", expand=True, side="left")
    scroll_mus = ttk.Scrollbar(musteri_table_container, orient="vertical", command=musteri_table.yview)
    musteri_table.configure(yscrollcommand=scroll_mus.set)
    scroll_mus.pack(side="right", fill="y")
    attach_context_delete(musteri_table)
    load_musteriler(musteri_table)
    
    # Zebra görünümü uygula
    apply_zebra_striping(musteri_table)

    # --- Müşteri Tablosu Arama ---
    def filter_musteriler(*args):
        query = search_var_musteri.get().lower()
        musteri_table.delete(*musteri_table.get_children())
        try:
            with open(MUSTERI_DOSYA, "r", encoding="utf-8") as f:
                data = json.load(f)
            for values in data:
                if any(query in str(v).lower() for v in values):
                    musteri_table.insert("", "end", values=values)
        except:
            pass
    search_var_musteri.trace("w", filter_musteriler)

    # --- Müşteri Şubelerini Göster ---
    def show_subeler(event):
        selected = musteri_table.selection()
        if not selected:
            return
        values = musteri_table.item(selected[0], "values")
        if not values or len(values) < 9:
            return

        subeler_str = values[8]
        subeler_list = [s.strip() for s in subeler_str.split(",") if s.strip()]

        win = tk.Toplevel(root)
        win.title(f"{values[3]} - Şubeler")
        win.geometry("400x300")

        tk.Label(win, text=f"Müşteri: {values[3]}").pack(anchor="w", padx=10, pady=5)

        listbox = tk.Listbox(win, height=10)
        listbox.pack(fill="both", expand=True, padx=10, pady=10)

        for s in subeler_list:
            listbox.insert(tk.END, s)

    musteri_table.bind("<Double-1>", show_subeler)

    # --- Müşteri Düzenleme ---
    def duzenle_musteri():
        selected = musteri_table.selection()
        if not selected:
            tk.messagebox.showwarning("Uyarı", "Lütfen düzenlemek için bir müşteri seçin!")
            return

        values = musteri_table.item(selected[0], "values")
        if not values:
            return

        win = tk.Toplevel(root)
        win.title("Müşteri Düzenle")
        win.geometry("600x500")

        labels = [
            "VKN/TCKN", "Adı", "Soyadı", "Unvan", "Vergi D. Şehir",
            "Vergi Dairesi", "Adres Şehir", "İlçe", "Şubeler", "Adres"
        ]
        entries = []

        for i, label in enumerate(labels):
            tk.Label(win, text=label + ":").grid(row=i, column=0, sticky="e", padx=5, pady=5)
            if label == "Adres":
                entry = tk.Text(win, width=50, height=4)
                entry.insert("1.0", values[i])
                entry.grid(row=i, column=1, padx=5, pady=5)
            else:
                entry = tk.Entry(win, width=50)
                entry.insert(0, values[i])
                entry.grid(row=i, column=1, padx=5, pady=5)
            entries.append(entry)

        def save_changes():
            new_values = []
            for i, label in enumerate(labels):
                if label == "Adres":
                    new_values.append(entries[i].get("1.0", "end").strip())
                else:
                    new_values.append(entries[i].get())
            
            # Güvenli güncelleme kullan (sadece bu müşteriyi güncelle)
            vkn = new_values[0]  # VKN ilk alanda
            if safe_update_musteri(vkn, new_values):
                # Tabloyu da güncelle
                musteri_table.item(selected[0], values=new_values)
                tk.messagebox.showinfo("Başarılı", "Müşteri bilgileri güncellendi")
            else:
                tk.messagebox.showerror("Hata", "Müşteri güncellenemedi")
            win.destroy()

        tk.Button(win, text="Kaydet", command=save_changes).grid(row=len(labels), column=1, pady=10)

    # --- Müşteri Silme ---
    def sil_musteri():
        selected = musteri_table.selection()
        if not selected:
            tk.messagebox.showwarning("Uyarı", "Lütfen silmek için bir müşteri seçin!")
            return

        answer = tk.messagebox.askyesno("Onay", "Bu müşteriyi silmek istediğinize emin misiniz?")
        if answer:
            # Her seçili müşteriyi güvenli şekilde sil
            deleted_count = 0
            for item in selected:
                values = musteri_table.item(item, "values")
                if values and len(values) > 0:
                    vkn = values[0]
                    if safe_delete_musteri(vkn):
                        musteri_table.delete(item)
                        deleted_count += 1
            
            if deleted_count > 0:
                tk.messagebox.showinfo("Başarılı", f"{deleted_count} müşteri silindi")
            else:
                tk.messagebox.showerror("Hata", "Hiçbir müşteri silinemedi")

    # Düzenle ve Sil butonları
    btn_frame = tk.Frame(frame_musteriler, bg="#d0d0d0")
    btn_frame.pack(pady=5)

    tk.Button(btn_frame, text="Düzenle", command=duzenle_musteri).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Sil", command=sil_musteri).pack(side="left", padx=5)
    # --- Müşteriler Sekmesi Sonu ---

    # ================== START ZİRVE BİLGİLERİ & LOG ==================
    frame_zirve_log = tk.Frame(notebook, padx=10, pady=10, bg="#d0d0d0")
    notebook.add(frame_zirve_log, text="Zirve Bilgileri & Log")

    # --- Zirve giriş bilgileri ---
    frame_zirve = tk.LabelFrame(frame_zirve_log, text="Zirve Giriş", padx=10, pady=10, bg="#d0d0d0")
    frame_zirve.pack(fill="x", pady=10)

    tk.Label(frame_zirve, text="Şirket:").grid(row=0, column=0, sticky="e")
    zirve_sirket_combo = ttk.Combobox(frame_zirve, values=[], width=25)
    zirve_sirket_combo.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(frame_zirve, text="Kullanıcı:").grid(row=0, column=2, sticky="e")
    zirve_user = tk.Entry(frame_zirve, width=25)
    zirve_user.grid(row=0, column=3, padx=5, pady=5)

    tk.Label(frame_zirve, text="Şifre:").grid(row=0, column=4, sticky="e")
    zirve_pass = tk.Entry(frame_zirve, width=25, show="*")
    zirve_pass.grid(row=0, column=5, padx=5, pady=5)

    # --- Kaydet & Sil butonları ---
    def kaydet_zirve():
        sirket = zirve_sirket_combo.get().strip()
        kullanici = zirve_user.get().strip()
        sifre = zirve_pass.get().strip()

        if not sirket:
            messagebox.showerror("Hata", "Lütfen şirket adını girin!")
            return

        bilgiler = {}
        if os.path.exists("zirve_bilgileri.json"):
            try:
                with open("zirve_bilgileri.json", "r", encoding="utf-8") as f:
                    bilgiler = json.load(f)
            except:
                pass

        bilgiler[sirket] = {"kullanici": kullanici, "sifre": sifre}

        with open("zirve_bilgileri.json", "w", encoding="utf-8") as f:
            json.dump(bilgiler, f, ensure_ascii=False, indent=4)

        zirve_sirket_combo["values"] = list(bilgiler.keys())
        messagebox.showinfo("Bilgi", f"{sirket} için bilgiler kaydedildi.")

    def sil_zirve():
        sirket = zirve_sirket_combo.get().strip()
        if not sirket or not os.path.exists("zirve_bilgileri.json"):
            return

        with open("zirve_bilgileri.json", "r", encoding="utf-8") as f:
            bilgiler = json.load(f)

        if sirket in bilgiler:
            del bilgiler[sirket]

        with open("zirve_bilgileri.json", "w", encoding="utf-8") as f:
            json.dump(bilgiler, f, ensure_ascii=False, indent=4)

        zirve_user.delete(0, "end")
        zirve_pass.delete(0, "end")
        zirve_sirket_combo["values"] = list(bilgiler.keys())
        messagebox.showinfo("Bilgi", f"{sirket} için bilgiler silindi.")

    # --- Şirket seçildiğinde bilgileri yükle ---
    def sirket_secildi(event=None):
        secilen = zirve_sirket_combo.get().strip()
        if not secilen or not os.path.exists("zirve_bilgileri.json"):
            return
        try:
            with open("zirve_bilgileri.json", "r", encoding="utf-8") as f:
                bilgiler = json.load(f)
            if secilen in bilgiler and isinstance(bilgiler[secilen], dict):
                zirve_user.delete(0, "end")
                zirve_pass.delete(0, "end")
                zirve_user.insert(0, bilgiler[secilen].get("kullanici", ""))
                zirve_pass.insert(0, bilgiler[secilen].get("sifre", ""))
        except Exception as e:
            print("⚠️ Şirket seçilirken hata:", e)

    zirve_sirket_combo.bind("<<ComboboxSelected>>", sirket_secildi)

    tk.Button(frame_zirve, text="Kaydet", command=kaydet_zirve).grid(row=0, column=6, padx=5, pady=5)
    tk.Button(frame_zirve, text="Sil", command=sil_zirve).grid(row=0, column=7, padx=5, pady=5)

    # Headless mod (Sil butonunun yanında)
    headless_var = tk.BooleanVar(value=False)
    headless_checkbox = tk.Checkbutton(frame_zirve, text="Headless", variable=headless_var, bg="#d0d0d0")
    headless_checkbox.grid(row=0, column=8, padx=5, pady=5)

    # --- Program açıldığında kayıtlı bilgileri yükle ---
    bilgiler = {}
    if os.path.exists("zirve_bilgileri.json"):
        try:
            with open("zirve_bilgileri.json", "r", encoding="utf-8") as f:
                bilgiler = json.load(f)
        except:
            pass
    else:
        bilgiler = {
            "Şirket A": {"kullanici": "kullanici_adi_A", "sifre": "sifreA123"},
            "Şirket B": {"kullanici": "kullanici_adi_B", "sifre": "sifreB456"}
        }
        with open("zirve_bilgileri.json", "w", encoding="utf-8") as f:
            json.dump(bilgiler, f, ensure_ascii=False, indent=4)

    if bilgiler:
        sirketler = list(bilgiler.keys())
        zirve_sirket_combo["values"] = sirketler
        ilk = sirketler[0]
        zirve_sirket_combo.set(ilk)

        if isinstance(bilgiler[ilk], dict):  # ✅ güvenlik kontrolü eklendi
            zirve_user.insert(0, bilgiler[ilk].get("kullanici", ""))
            zirve_pass.insert(0, bilgiler[ilk].get("sifre", ""))
        else:
            print("⚠️ Beklenmedik JSON formatı:", type(bilgiler[ilk]))
    else:
        zirve_sirket_combo["values"] = ["Şirket A", "Şirket B"]

    # --- Log ekranı ---
    frame_log = tk.LabelFrame(frame_zirve_log, text="İşlem Logu", padx=10, pady=10, bg="#d0d0d0")
    frame_log.pack(fill="both", expand=True, pady=10)

    log_text = tk.Text(frame_log, state="disabled", height=15)
    log_text.pack(fill="both", expand=True)

    # 👇 Kuyruk görüntüleme tablosunu başlat
    init_queue_view(frame_zirve_log)
    # ================== END ZİRVE BİLGİLERİ & LOG ==================




# ================== START MAIN SCRIPT ==================
driver_global = None  # 👈 Chrome'u global tanımladık
fatura_queue = []     # 👈 Fatura kuyruğu burada tanımlı
tamamlanan_faturalar = []  # 👈 Tamamlanan faturalar listesi
is_processing = False # 👈 Şu an işlem var mı?
headless_var = None   # 👈 Headless seçeneği (GUI içinde ayarlanır)

# ================== END MAIN SCRIPT ==================

# ================== START FATURA OKUMA FONKSİYONU ==================
def read_invoices_from_zirve():
    """Zirve portalından E-Fatura ve E-Arşiv faturalarını okur"""
    import threading
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium.webdriver.chrome.options import Options
    
    def read_invoices():
        try:
            log_yaz("🔍 Faturalar okunuyor...")
            
            # Chrome başlat
            service = Service(ChromeDriverManager().install())
            chrome_options = Options()
            if headless_var is not None and headless_var.get():
                # Headless mod için ek seçenekler
                chrome_options.add_argument("--headless=new")
                chrome_options.add_argument("--window-size=1920,1080")
                chrome_options.add_argument("--disable-gpu")
                chrome_options.add_argument("--disable-dev-shm-usage")
                chrome_options.add_argument("--disable-extensions")
                chrome_options.add_argument("--no-sandbox")
                chrome_options.add_argument("--disable-web-security")
                chrome_options.add_argument("--disable-features=VizDisplayCompositor")
                chrome_options.add_argument("--remote-debugging-port=9222")
            else:
                # Normal mod
                chrome_options.add_argument("--start-maximized")
                chrome_options.add_argument("--disable-web-security")
                chrome_options.add_argument("--disable-features=VizDisplayCompositor")
                chrome_options.add_argument("--disable-extensions")
                chrome_options.add_argument("--no-sandbox")
            
            driver = webdriver.Chrome(service=service, options=chrome_options)
            if not (headless_var is not None and headless_var.get()):
                driver.maximize_window()
            
            # Zirve portalına giriş
            driver.get("https://yeniportal.zirvedonusum.com/accounting/login")
            
            # Giriş bilgileri
            username = zirve_user.get().strip()
            password = zirve_pass.get().strip()
            
            if not (username and password):
                log_yaz("❌ Zirve giriş bilgileri eksik!")
                return
            
            # Giriş yap
            username_field = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.NAME, "username"))
            )
            password_field = driver.find_element(By.NAME, "password")
            
            username_field.send_keys(username)
            password_field.send_keys(password)
            
            # Giriş butonuna tıkla
            try:
                login_btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(text(),'GİRİŞ')]"))
                )
                login_btn.click()
            except:
                from selenium.webdriver.common.keys import Keys
                password_field.send_keys(Keys.RETURN)
            
            # Giriş kontrolü - e-Dönüşüm menüsünün yüklenmesini bekle (Headless modda daha uzun bekle)
            wait_time = 30 if (headless_var is not None and headless_var.get()) else 20
            try:
                WebDriverWait(driver, wait_time).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@href='#pagesTransformation']"))
                )
                log_yaz("✅ Portal giriş başarılı, e-Dönüşüm menüsü hazır!")
            except:
                log_yaz("⚠️ Giriş kontrolü yapılamadı, devam ediliyor...")
            
            # e-Dönüşüm menüsüne tıkla (zaten giriş kontrolünde tıklanabilir hale geldi)
            try:
                edonusum_menu = driver.find_element(By.XPATH, "//a[@href='#pagesTransformation']")
                edonusum_menu.click()
                log_yaz("✅ e-Dönüşüm menüsüne tıklandı")
            except Exception as e:
                log_yaz(f"⚠️ e-Dönüşüm menüsü bulunamadı: {e}")
            
            # E-Fatura faturalarını oku
            try:
                log_yaz("🔍 E-Fatura faturaları okunuyor...")
                
                # e-Fatura menüsüne tıkla
                efatura_menu = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@data-toggle='collapse' and @href='#eInvoice']"))
                )
                efatura_menu.click()
                log_yaz("✅ e-Fatura menüsüne tıklandı")
                
                # Giden Faturalar linkine tıkla
                giden_faturalar_link = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//span[contains(text(),'Giden Faturalar')]"))
                )
                giden_faturalar_link.click()
                log_yaz("✅ Giden Faturalar linkine tıklandı")
                
                # Sayfa yüklenmesini bekle
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.TAG_NAME, "table"))
                )
                
                # Tabloları bul
                tables = WebDriverWait(driver, 20).until(
                    EC.presence_of_all_elements_located((By.TAG_NAME, "table"))
                )
                
                # Doğru tabloyu bul
                hedef_tablo = None
                for i, t in enumerate(tables):
                    try:
                        header = t.find_element(By.TAG_NAME, "thead").text
                        if "Fatura No" in header and "VKN" in header and "Fatura Tarihi" in header:
                            hedef_tablo = t
                            log_yaz(f"✅ E-Fatura tablosu bulundu: Tablo {i}")
                            break
                    except:
                        continue
                
                if not hedef_tablo:
                    log_yaz("❌ E-Fatura tablosu bulunamadı")
                else:
                    # E-Fatura tablosunu temizle
                    for item in efatura_table.get_children():
                        efatura_table.delete(item)
                    
                    # Satırları oku
                    rows = hedef_tablo.find_elements(By.TAG_NAME, "tr")[1:]  # Başlık satırını atla
                    efatura_count = 0
                    
                    for i, row in enumerate(rows):
                        cells = row.find_elements(By.TAG_NAME, "td")
                        if len(cells) < 10:
                            continue
                        
                        try:
                            # Verileri oku (doğru sütun indeksleri)
                            # Sütun 3: Alıcı Unvan + VKN/TCKN (aynı hücrede, <br> ile ayrılmış)
                            unvan_vkn_text = cells[3].text.strip() if len(cells) > 3 else ""
                            lines = [l.strip() for l in unvan_vkn_text.split("\n") if l.strip()]
                            musteri = lines[0] if len(lines) > 0 else ""
                            vergi_no = lines[1] if len(lines) > 1 else ""
                            
                            # Sütun 2: Fatura Tarihi + Alınma Tarihi
                            tarih_text = cells[2].text.strip() if len(cells) > 2 else ""
                            tarih_lines = [l.strip() for l in tarih_text.split("\n") if l.strip()]
                            tarih = tarih_lines[0] if len(tarih_lines) > 0 else ""
                            
                            # Sütun 6: Ödenecek Tutar + VHTT
                            tutar_text = cells[6].text.strip() if len(cells) > 6 else ""
                            tutar_lines = [l.strip() for l in tutar_text.split("\n") if l.strip()]
                            tutar = tutar_lines[0] if len(tutar_lines) > 0 else ""
                            
                            # Sütun 1: Fatura No + ETTN
                            fatura_no_text = cells[1].text.strip() if len(cells) > 1 else ""
                            fatura_lines = [l.strip() for l in fatura_no_text.split("\n") if l.strip()]
                            fatura_no = fatura_lines[0] if len(fatura_lines) > 0 else ""
                            
                            if musteri:
                                efatura_table.insert("", "end", values=(musteri, vergi_no, tutar, "E-Fatura", tarih, fatura_no))
                                efatura_count += 1
                                log_yaz(f"✅ E-Fatura okundu: {musteri} - {vergi_no} - {tutar}")
                        except Exception as e:
                            log_yaz(f"⚠️ E-Fatura satırı okunamadı: {e}")
                            continue
                    
                    log_yaz(f"📋 {efatura_count} adet E-Fatura okundu")
                    
            except Exception as e:
                log_yaz(f"❌ E-Fatura okuma hatası: {e}")
            
            # E-Arşiv faturalarını oku
            try:
                log_yaz("🔍 E-Arşiv faturaları okunuyor...")
                
                # e-Arşiv menüsüne tıkla
                earsiv_menu = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//p[contains(text(),'e-Arşiv')]"))
                )
                earsiv_menu.click()
                log_yaz("✅ e-Arşiv menüsüne tıklandı")
                
                # e-Arşiv Faturalar linkine tıkla
                earsiv_faturalar_link = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//span[contains(text(),'e-Arşiv Faturalar')]"))
                )
                earsiv_faturalar_link.click()
                log_yaz("✅ e-Arşiv Faturalar linkine tıklandı")
                
                # Sayfa yüklenmesini bekle
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.TAG_NAME, "table"))
                )
                
                # Tabloları bul
                tables = WebDriverWait(driver, 20).until(
                    EC.presence_of_all_elements_located((By.TAG_NAME, "table"))
                )
                
                # Doğru tabloyu bul
                hedef_tablo = None
                for i, t in enumerate(tables):
                    try:
                        header = t.find_element(By.TAG_NAME, "thead").text
                        if "Doküman No" in header and "VKN" in header and "Fatura Tarihi" in header:
                            hedef_tablo = t
                            log_yaz(f"✅ E-Arşiv tablosu bulundu: Tablo {i}")
                            break
                    except:
                        continue
                
                if not hedef_tablo:
                    log_yaz("❌ E-Arşiv tablosu bulunamadı")
                else:
                    # E-Arşiv tablosunu temizle
                    for item in earsiv_table.get_children():
                        earsiv_table.delete(item)
                    
                    # Satırları oku
                    rows = hedef_tablo.find_elements(By.TAG_NAME, "tr")[1:]  # Başlık satırını atla
                    earsiv_count = 0
                    
                    for i, row in enumerate(rows):
                        cells = row.find_elements(By.TAG_NAME, "td")
                        if len(cells) < 11:
                            continue
                        
                        try:
                            # Verileri oku (doğru sütun indeksleri)
                            # Sütun 3: Alıcı Ünvan + VKN/TCKN (aynı hücrede, <br> ile ayrılmış)
                            unvan_vkn_text = cells[3].text.strip() if len(cells) > 3 else ""
                            lines = [l.strip() for l in unvan_vkn_text.split("\n") if l.strip()]
                            musteri = lines[0] if len(lines) > 0 else ""
                            vergi_no = lines[1] if len(lines) > 1 else ""
                            
                            # Sütun 2: Fatura Tarihi + Alınma Tarihi
                            tarih_text = cells[2].text.strip() if len(cells) > 2 else ""
                            tarih_lines = [l.strip() for l in tarih_text.split("\n") if l.strip()]
                            tarih = tarih_lines[0] if len(tarih_lines) > 0 else ""
                            
                            # Sütun 6: Ödenecek Tutar + VHTT
                            tutar_text = cells[6].text.strip() if len(cells) > 6 else ""
                            tutar_lines = [l.strip() for l in tutar_text.split("\n") if l.strip()]
                            tutar = tutar_lines[0] if len(tutar_lines) > 0 else ""
                            
                            # Sütun 1: Doküman No + ETTN
                            fatura_no_text = cells[1].text.strip() if len(cells) > 1 else ""
                            fatura_lines = [l.strip() for l in fatura_no_text.split("\n") if l.strip()]
                            fatura_no = fatura_lines[0] if len(fatura_lines) > 0 else ""
                            
                            if musteri:
                                earsiv_table.insert("", "end", values=(musteri, vergi_no, tutar, "E-Arşiv", tarih, fatura_no))
                                earsiv_count += 1
                                log_yaz(f"✅ E-Arşiv okundu: {musteri} - {vergi_no} - {tutar}")
                        except Exception as e:
                            log_yaz(f"⚠️ E-Arşiv satırı okunamadı: {e}")
                            continue
                    
                    log_yaz(f"📋 {earsiv_count} adet E-Arşiv okundu")
                    
            except Exception as e:
                log_yaz(f"❌ E-Arşiv okuma hatası: {e}")
            
            driver.quit()
            log_yaz("✅ Fatura okuma tamamlandı")
            
        except Exception as e:
            log_yaz(f"❌ Fatura okuma hatası: {e}")
    
    threading.Thread(target=read_invoices).start()


# ================== END FATURA OKUMA FONKSİYONU ==================








def guncelle_subeler():
    """Seçilen faturalara göre şubeleri güncelle"""
    try:
        # E-Fatura tablosundan seçilenleri al
        efatura_selected = efatura_table.selection()
        earsiv_selected = earsiv_table.selection()
        
        if not efatura_selected and not earsiv_selected:
            return
        
        # Müşteri verilerini oku
        try:
            with open("musteriler.json", "r", encoding="utf-8") as f:
                musteri_verileri = json.load(f)
        except Exception as e:
            log_yaz(f"❌ Müşteri verileri okunamadı: {e}")
            return
        
        # Seçilen faturalardaki VKN'leri topla
        global secilen_musteri_vknleri
        secilen_vknler = set()
        
        # E-Fatura seçilenlerini işle
        for item in efatura_selected:
            values = efatura_table.item(item, "values")
            if len(values) > 1:  # VKN sütunu (index 1)
                vkn = values[1].strip()
                if vkn:
                    secilen_vknler.add(vkn)
                    log_yaz(f"📋 E-Fatura VKN seçildi: {vkn}")
        
        # E-Arşiv seçilenlerini işle
        for item in earsiv_selected:
            values = earsiv_table.item(item, "values")
            if len(values) > 1:  # VKN sütunu (index 1)
                vkn = values[1].strip()
                if vkn:
                    secilen_vknler.add(vkn)
                    log_yaz(f"📋 E-Arşiv VKN seçildi: {vkn}")
        
        # Global değişkeni güncelle
        secilen_musteri_vknleri = secilen_vknler
        
        # VKN'ler ile eşleşen müşterilerin şubelerini bul
        eslesen_subeler = set()
        
        for musteri in musteri_verileri:
            if len(musteri) >= 10:  # Yeterli veri var mı kontrol et
                vkn = musteri[0].strip()
                if vkn in secilen_vknler:
                    # Sadece şube bilgilerini al (index 8)
                    subeler = []
                    if musteri[8]:  # a,b,c,d
                        subeler.extend([s.strip() for s in musteri[8].split(",") if s.strip()])
                    
                    for sube in subeler:
                        if sube:
                            eslesen_subeler.add(sube)
                    
                    log_yaz(f"✅ VKN {vkn} eşleşti: {musteri[3] if len(musteri) > 3 else 'Bilinmeyen'}")
        
        # Şube combobox'ını güncelle
        if eslesen_subeler:
            subeler_listesi = sorted(list(eslesen_subeler))
            fatura_kes_sube_combo['values'] = subeler_listesi
            fatura_kes_sube_combo.set(subeler_listesi[0])
            log_yaz(f"🏢 {len(subeler_listesi)} şube bulundu: {', '.join(subeler_listesi)}")
        else:
            # Şube bulunamadığında combobox'ı temizle
            fatura_kes_sube_combo['values'] = []
            fatura_kes_sube_combo.set("")
            log_yaz("⚠️ Seçilen faturalar için şube bulunamadı - combobox temizlendi")
        
    except Exception as e:
        log_yaz(f"❌ Şube güncelleme hatası: {e}")

def process_fatura_indirme_kuyrugu():
    """Fatura indirme kuyruğunu işler - Chrome'u bir kez açıp tüm faturaları indirir"""
    global fatura_indirme_aktif, fatura_indirme_kuyrugu, zirve_user, zirve_pass
    
    if not fatura_indirme_kuyrugu:
        return
    
    fatura_indirme_aktif = True
    toplam_kuyruk_sayisi = len(fatura_indirme_kuyrugu)
    log_yaz(f"🚀 Fatura indirme kuyruğu başlatılıyor... (Toplam: {toplam_kuyruk_sayisi} işlem)")
    
    # Chrome driver'ı bir kez başlat
    driver = None
    try:
        # Chrome driver'ı başlat
        service = Service(ChromeDriverManager().install())
        options = webdriver.ChromeOptions()
        
        options.add_argument("--start-maximized")
        options.add_argument("--disable-web-security")
        options.add_argument("--disable-features=VizDisplayCompositor")
        options.add_argument("--disable-extensions")
        options.add_argument("--no-sandbox")
        
        # İndirme klasörünü ayarla
        download_dir = os.path.join(os.getcwd(), "indirilen_faturalar")
        if not os.path.exists(download_dir):
            os.makedirs(download_dir)
        
        prefs = {
            "download.default_directory": download_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
            "profile.default_content_settings.popups": 0,
            "profile.default_content_setting_values.automatic_downloads": 1
        }
        options.add_experimental_option("prefs", prefs)
        
        driver = webdriver.Chrome(service=service, options=options)
        driver.maximize_window()
        
        # Zirve portalına giriş yap (bir kez)
        log_yaz("🔐 Zirve portalına giriş yapılıyor...")
        driver.get("https://yeniportal.zirvedonusum.com/accounting/login")
        
        kullanici = zirve_user.get().strip()
        sifre = zirve_pass.get().strip()
        
        if not kullanici or not sifre:
            log_yaz("❌ Kullanıcı adı veya şifre boş!")
            return
        
        # Portal giriş
        login_portal(driver, kullanici, sifre)
        log_yaz("✅ Portal giriş başarılı!")
        
        # Toplam indirilen fatura sayısı
        toplam_indirilen = 0
        
        # Kuyruktaki her işlemi sırayla işle
        while fatura_indirme_kuyrugu:
            kuyruk_item = fatura_indirme_kuyrugu.pop(0)
            
            log_yaz(f"📥 Kuyruktan işlem alınıyor... (Kalan: {len(fatura_indirme_kuyrugu)})")
            
            try:
                # Kuyruk item'ından verileri al
                efatura_selected = kuyruk_item['efatura_selected']
                earsiv_selected = kuyruk_item['earsiv_selected']
                sube_degeri = kuyruk_item['sube_degeri']
                personel_degeri = kuyruk_item['personel_degeri']
                islem_turu_degeri = kuyruk_item['islem_turu_degeri']
                
                # Bu işlem için fatura indirme
                indirilen_sayi = fatura_indir_session(driver, efatura_selected, earsiv_selected, 
                                                     sube_degeri, personel_degeri, islem_turu_degeri, download_dir)
                toplam_indirilen += indirilen_sayi
                
            except Exception as e:
                log_yaz(f"❌ Kuyruk işlemi hatası: {e}")
                continue
        
        log_yaz(f"🎉 Tüm kuyruk tamamlandı! Toplam {toplam_indirilen} fatura indirildi")
        
    except Exception as e:
        log_yaz(f"❌ Genel kuyruk hatası: {e}")
    finally:
        # Chrome'u kapat
        if driver:
            try:
                driver.quit()
                log_yaz("🔒 Chrome tarayıcısı kapatıldı")
            except:
                pass
        
        fatura_indirme_aktif = False
        log_yaz("✅ Fatura indirme kuyruğu tamamlandı")

def fatura_indir_session(driver, efatura_selected, earsiv_selected, sube_degeri, personel_degeri, islem_turu_degeri, download_dir):
    """Mevcut driver session'ını kullanarak faturaları indirir"""
    indirilen_sayisi = 0
    
    try:
        # E-Fatura seçilenlerini işle
        if efatura_selected:
            log_yaz("📄 E-Fatura sayfasına gidiliyor...")
            
            # E-Dönüşüm menüsüne tıkla
            try:
                e_donusum_menu = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@data-toggle='collapse' and @href='#pagesTransformation']"))
                )
                e_donusum_menu.click()
                log_yaz("✅ E-Dönüşüm menüsüne tıklandı")
            except Exception as e:
                log_yaz(f"❌ E-Dönüşüm menüsü bulunamadı: {e}")
                return indirilen_sayisi
            
            # E-Fatura menüsüne tıkla
            try:
                e_fatura_menu = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@data-toggle='collapse' and @href='#eInvoice']"))
                )
                e_fatura_menu.click()
                log_yaz("✅ E-Fatura menüsüne tıklandı")
            except Exception as e:
                log_yaz(f"❌ E-Fatura menüsü bulunamadı: {e}")
                return indirilen_sayisi
            
            # Giden Faturalar linkine tıkla
            try:
                giden_faturalar = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@href='/accounting/eInvoiceOutbox']"))
                )
                giden_faturalar.click()
                log_yaz("✅ Giden Faturalar linkine tıklandı")
            except Exception as e:
                log_yaz(f"❌ Giden Faturalar linki bulunamadı: {e}")
                return indirilen_sayisi
            
            # Tabloları bekle
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "table"))
            )
            
            # E-Fatura seçilenlerini işle
            for item in efatura_selected:
                try:
                    values = efatura_table.item(item, "values")
                    if len(values) < 6:
                        continue
                        
                    musteri_adi = values[0].strip()
                    vkn = values[1].strip()
                    fatura_no = values[5].strip()
                    
                    # VKN'ye göre müşteri ismini bul
                    musteri_unvani = vkn_ile_musteri_ismi_bul(vkn)
                    if musteri_unvani:
                        musteri_adi = musteri_unvani
                        log_yaz(f"📋 Müşteri unvanı bulundu: {musteri_adi}")
                    else:
                        log_yaz(f"⚠️ VKN {vkn} için unvan bulunamadı, mevcut isim kullanılıyor: {musteri_adi}")
                    
                    # Fatura isimlendirmesini oluştur
                    fatura_adi = musteri_adi
                    
                    if sube_degeri:
                        fatura_adi += f" - {sube_degeri}"
                    if personel_degeri:
                        fatura_adi += f" - {personel_degeri}"
                    if islem_turu_degeri:
                        fatura_adi += f" - {islem_turu_degeri}"
                    
                    fatura_adi += f" - {fatura_no}"
                    
                    log_yaz(f"📥 E-Fatura indiriliyor: {fatura_adi}")
                    
                    # Fatura indirme işlemi
                    if indir_fatura_from_table(driver, fatura_no, fatura_adi, download_dir):
                        indirilen_sayisi += 1
                        
                except Exception as e:
                    log_yaz(f"⚠️ E-Fatura indirme hatası: {e}")
                    continue
        
        # E-Arşiv seçilenlerini işle
        if earsiv_selected:
            log_yaz("📄 E-Arşiv sayfasına gidiliyor...")
            
            # E-Arşiv sayfasına git
            try:
                # E-Dönüşüm menüsüne tıkla
                e_donusum_menu = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@data-toggle='collapse' and @href='#pagesTransformation']"))
                )
                e_donusum_menu.click()
                
                # E-Arşiv menüsüne tıkla
                e_arsiv_menu = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@data-toggle='collapse' and @href='#eArchive']"))
                )
                e_arsiv_menu.click()
                
                # E-Arşiv Giden Faturalar linkine tıkla
                earsiv_giden_faturalar = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@href='/accounting/eArchiveOutbox']"))
                )
                earsiv_giden_faturalar.click()
                log_yaz("✅ E-Arşiv sayfasına gidildi")
                
                # Tabloları bekle
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, "table"))
                )
                
            except Exception as e:
                log_yaz(f"❌ E-Arşiv sayfasına gidilemedi: {e}")
                return indirilen_sayisi
            
            # E-Arşiv seçilenlerini işle
            for item in earsiv_selected:
                try:
                    values = earsiv_table.item(item, "values")
                    if len(values) < 6:
                        continue
                        
                    musteri_adi = values[0].strip()
                    vkn = values[1].strip()
                    fatura_no = values[5].strip()
                    
                    # VKN'ye göre müşteri ismini bul
                    musteri_unvani = vkn_ile_musteri_ismi_bul(vkn)
                    if musteri_unvani:
                        musteri_adi = musteri_unvani
                        log_yaz(f"📋 Müşteri unvanı bulundu: {musteri_adi}")
                    else:
                        log_yaz(f"⚠️ VKN {vkn} için unvan bulunamadı, mevcut isim kullanılıyor: {musteri_adi}")
                    
                    # Fatura isimlendirmesini oluştur
                    fatura_adi = musteri_adi
                    
                    if sube_degeri:
                        fatura_adi += f" - {sube_degeri}"
                    if personel_degeri:
                        fatura_adi += f" - {personel_degeri}"
                    if islem_turu_degeri:
                        fatura_adi += f" - {islem_turu_degeri}"
                    
                    fatura_adi += f" - {fatura_no}"
                    
                    log_yaz(f"📥 E-Arşiv indiriliyor: {fatura_adi}")
                    
                    # Fatura indirme işlemi
                    if indir_fatura_from_table(driver, fatura_no, fatura_adi, download_dir):
                        indirilen_sayisi += 1
                        
                except Exception as e:
                    log_yaz(f"⚠️ E-Arşiv indirme hatası: {e}")
                    continue
        
    except Exception as e:
        log_yaz(f"❌ Session fatura indirme hatası: {e}")
    
    return indirilen_sayisi

def indir_fatura_from_table(driver, fatura_no, fatura_adi, download_dir):
    """Tablodaki belirli bir faturayı indir (stale elementi tolere ederek ve modalı kapatarak)."""
    try:
        for attempt in range(3):
            try:
                log_yaz(f"🔍 Fatura numarası aranıyor: {fatura_no}")
                rows = driver.find_elements(By.TAG_NAME, "tr")
                hedef_row = None
                for row in rows:
                    tds = row.find_elements(By.TAG_NAME, "td")
                    if len(tds) < 1:
                        continue
                    for td in tds:
                        if fatura_no in td.text.strip():
                            hedef_row = row
                            break
                    if hedef_row is not None:
                        break
                if hedef_row is None:
                    return False

                # Dropdown aç
                dropdown = hedef_row.find_element(By.CSS_SELECTOR, "button[data-toggle='dropdown']")
                driver.execute_script("arguments[0].click();", dropdown)
                time.sleep(0.5)

                # Fatura PDF İndir'i tıkla
                WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "a.dropdown-item")))
                links = driver.find_elements(By.CSS_SELECTOR, "a.dropdown-item")
                pdf_link = None
                for a in links:
                    txt = a.text.strip()
                    if "Fatura PDF İndir" in txt or "PDF" in txt:
                        pdf_link = a
                        break
                if pdf_link is None:
                    return False
                pdf_link.click()
                time.sleep(1.5)

                # Yeni pencereye geç ve indir
                handles = driver.window_handles
                if len(handles) > 1:
                    driver.switch_to.window(handles[-1])
                try:
                    indir_btn = None
                    try:
                        indir_btn = WebDriverWait(driver, 4).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div#icon")))
                    except:
                        try:
                            indir_btn = WebDriverWait(driver, 4).until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(),'Pdf İndir')]")))
                        except:
                            indir_btn = WebDriverWait(driver, 4).until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(),'İndir')]")))
                    indir_btn.click()
                    time.sleep(2)

                    # Dosyayı benzersiz isimle kaydet
                    indirilenler = [f for f in os.listdir(download_dir) if f.endswith(".pdf")]
                    if indirilenler:
                        kaynak = max([os.path.join(download_dir, f) for f in indirilenler], key=os.path.getctime)
                        hedef = os.path.join(download_dir, f"{fatura_adi}.pdf")
                        if os.path.exists(hedef):
                            n = 1
                            while os.path.exists(os.path.join(download_dir, f"{fatura_adi} ({n}).pdf")):
                                n += 1
                            hedef = os.path.join(download_dir, f"{fatura_adi} ({n}).pdf")
                        os.rename(kaynak, hedef)
                        log_yaz(f"✅ Fatura indirildi: {os.path.basename(hedef)}")
                finally:
                    # Yeni pencereyi kapat ve ana pencereye dön
                    if len(handles) > 1:
                        driver.close()
                        driver.switch_to.window(handles[0])

                # Modalı kapat (id=close ya da class=close)
                try:
                    btn = driver.find_element(By.CSS_SELECTOR, "button#close")
                    btn.click()
                    try:
                        WebDriverWait(driver, 3).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.modal")))
                    except:
                        time.sleep(0.3)
                except:
                    try:
                        btn2 = driver.find_element(By.XPATH, "//button[@class='close pull-right' and @title='Kapat']")
                        btn2.click()
                        try:
                            WebDriverWait(driver, 3).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.modal")))
                        except:
                            time.sleep(0.3)
                    except:
                        pass

                return True
            except StaleElementReferenceException:
                log_yaz("Stale element, yeniden deniyorum...")
                time.sleep(0.5)
                continue
        return False
    except Exception as e:
        log_yaz(f"❌ Fatura arama hatası: {e}")
        return False

# Eski fatura indirme fonksiyonu kaldırıldı - Artık optimize edilmiş kuyruk sistemi kullanılıyor
# Chrome'u bir kez açıp tüm faturaları sırayla indiren process_fatura_indirme_kuyrugu() fonksiyonu kullanılıyor

def indir_secilen_faturalar():
    """Seçilen faturaları indir"""
    global fatura_kes_sube_combo, fatura_kes_personel_entry, fatura_kes_islem_turu_combo, fatura_indirme_aktif, fatura_indirme_kuyrugu
    
    try:
        # E-Fatura tablosundan seçilenleri al
        efatura_selected = efatura_table.selection()
        earsiv_selected = earsiv_table.selection()
        
        if not efatura_selected and not earsiv_selected:
            log_yaz("⚠️ Lütfen indirmek istediğiniz faturaları seçin")
            return
        
        log_yaz(f"🔍 {len(efatura_selected)} E-Fatura, {len(earsiv_selected)} E-Arşiv seçildi")
        
        # O anki GUI değerlerini yakala
        sube_degeri = fatura_kes_sube_combo.get().strip()
        personel_degeri = fatura_kes_personel_entry.get().strip()
        islem_turu_degeri = fatura_kes_islem_turu_combo.get().strip()
        
        # Seçilen faturaları kuyruğa ekle
        fatura_indirme_kuyrugu.append({
            'efatura_selected': efatura_selected,
            'earsiv_selected': earsiv_selected,
            'sube_degeri': sube_degeri,
            'personel_degeri': personel_degeri,
            'islem_turu_degeri': islem_turu_degeri,
            'timestamp': time.time()
        })
        
        log_yaz(f"📋 Fatura indirme kuyruğa eklendi (Personel: {personel_degeri}, İşlem: {islem_turu_degeri}). Kuyruk sırası: {len(fatura_indirme_kuyrugu)}")
        
        # Eğer şu an işlem yapılmıyorsa kuyruğu başlat
        if not fatura_indirme_aktif:
            threading.Thread(target=process_fatura_indirme_kuyrugu, daemon=True).start()
        
    except Exception as e:
        log_yaz(f"❌ Fatura indirme hatası: {e}")

# ================== START ŞUBE YÖNETİMİ FONKSİYONLARI ==================

def sube_musterilere_ekle(yeni_sube):
    """Seçilen müşterilerin şubeler alanına (index 8) yeni şubeyi ekle"""
    global secilen_musteri_vknleri
    
    try:
        if not secilen_musteri_vknleri:
            log_yaz("⚠️ Önce fatura seçiniz!")
            return
        
        # musteriler.json'u oku
        if not os.path.exists("musteriler.json"):
            log_yaz("⚠️ musteriler.json dosyası bulunamadı")
            return
        
        with open("musteriler.json", "r", encoding="utf-8") as f:
            musteriler = json.load(f)
        
        guncellenen_sayisi = 0
        
        for musteri in musteriler:
            if len(musteri) > 0:
                vkn = str(musteri[0]).strip().lstrip('0') or '0'
                # Seçilen VKN'leri de normalize et
                normalized_secilen = {str(v).strip().lstrip('0') or '0' for v in secilen_musteri_vknleri}
                if vkn in normalized_secilen:
                    # Müşteri unvanını al
                    unvan = musteri[3] if len(musteri) > 3 else "Bilinmeyen Müşteri"
                    
                    # Mevcut şubeleri kontrol et (index 8)
                    mevcut_subeler = []
                    if len(musteri) > 8 and musteri[8]:
                        mevcut_subeler = [s.strip() for s in musteri[8].split(',') if s.strip()]
                    
                    # Yeni şube zaten yoksa ekle
                    if yeni_sube not in mevcut_subeler:
                        mevcut_subeler.append(yeni_sube)
                        # Şubeler alanını güncelle (virgülle ayırarak)
                        if len(musteri) > 8:
                            musteri[8] = ','.join(mevcut_subeler)
                        else:
                            # Eğer liste yeterince uzun değilse, eksik alanları boş string ile doldur
                            while len(musteri) < 9:
                                musteri.append("")
                            musteri[8] = yeni_sube
                        
                        guncellenen_sayisi += 1
                        log_yaz(f"✅ '{unvan}' müşterisine '{yeni_sube}' şubesi eklendi")
                    else:
                        log_yaz(f"⚠️ '{unvan}' müşterisinde '{yeni_sube}' şubesi zaten mevcut")
        
        if guncellenen_sayisi > 0:
            # Dosyayı kaydet
            with open("musteriler.json", "w", encoding="utf-8") as f:
                json.dump(musteriler, f, ensure_ascii=False, indent=4)
            
            log_yaz(f"✅ {guncellenen_sayisi} müşteriye '{yeni_sube}' şubesi eklendi")
            
            # Şube combobox'ını yeniden güncelle
            guncelle_sube_combobox()
        else:
            log_yaz("⚠️ Hiçbir müşteriye yeni şube eklenemedi")
            
    except Exception as e:
        log_yaz(f"❌ Şube ekleme hatası: {e}")

def guncelle_sube_combobox():
    """Seçilen müşterilere göre şube combobox'ını güncelle"""
    global secilen_musteri_vknleri, fatura_kes_sube_combo
    
    try:
        if not os.path.exists("musteriler.json"):
            return
        
        with open("musteriler.json", "r", encoding="utf-8") as f:
            musteriler = json.load(f)
        
        # Seçilen müşterilerin şubelerini topla
        eslesen_subeler = set()
        for musteri in musteriler:
            if len(musteri) > 0:
                vkn = str(musteri[0]).strip().lstrip('0') or '0'
                # Seçilen VKN'leri de normalize et
                normalized_secilen = {str(v).strip().lstrip('0') or '0' for v in secilen_musteri_vknleri}
                if vkn in normalized_secilen and len(musteri) > 8:
                    if musteri[8]:
                        subeler_listesi = [s.strip() for s in musteri[8].split(',') if s.strip()]
                        for sube in subeler_listesi:
                            eslesen_subeler.add(sube)
        
        # Combobox'ı güncelle
        if eslesen_subeler:
            subeler_listesi = sorted(list(eslesen_subeler))
            fatura_kes_sube_combo['values'] = subeler_listesi
            # Mevcut değer geçerli değilse ilkini seç
            if fatura_kes_sube_combo.get() not in subeler_listesi:
                fatura_kes_sube_combo.set(subeler_listesi[0])
            log_yaz(f"🔄 Şube listesi güncellendi: {', '.join(subeler_listesi)}")
        
    except Exception as e:
        log_yaz(f"❌ Şube combobox güncelleme hatası: {e}")

# ================== END ŞUBE YÖNETİMİ FONKSİYONLARI ==================

print("✅ GUI dosyası çalışıyor")
print("🔄 GitHub güncelleme kontrolü - 2025-09-20 17:15:00")

if __name__ == "__main__":
    gui_main()
    tk.mainloop()

