import tkinter as tk
from tkinter import ttk, messagebox
import os
import json
import threading
import traceback
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options

# Selenium fonksiyonlarını import et
from selenium_taslak_oluşturuyor import (
    login_portal,
    create_invoice_simple,
    check_customer_and_edit,
    upload_products_from_excel,
    add_invoice_note,
    save_and_close_invoice
)

# Global değişkenler
driver_global = None
fatura_queue = []
is_processing = False
headless_var = None
log_text = None
queue_table = None







# ================== START HELPERS ==================
class AutocompleteCombobox(tk.Frame):
    def __init__(self, master, values=None, width=20, next_widget=None,
                 linked_fields=None, kart_table=None, **kwargs):
        super().__init__(master, **kwargs)
        self.values = values if values else []
        self.next_widget = next_widget
        self.linked_fields = linked_fields
        self.kart_table = kart_table

        self.var = tk.StringVar()

        # Üst satır: Entry + ▼ butonu
        top_frame = tk.Frame(self)
        top_frame.pack(fill="x")

        self.entry = tk.Entry(top_frame, textvariable=self.var, width=width)
        self.entry.pack(side="left", fill="x", expand=True)

        self.button = tk.Button(top_frame, text="▼", width=2, command=self.show_all)
        self.button.pack(side="right")

        # Aşağı açılan liste
        self.listbox = tk.Listbox(self, height=5)
        self.listbox.bind("<<ListboxSelect>>", self.on_select)
        self.listbox.pack_forget()

        # Entry eventleri
        self.entry.bind("<KeyRelease>", self.on_keyrelease)
        self.entry.bind("<Down>", self.focus_listbox)
        self.entry.bind("<Return>", self.confirm_selection)
        self.entry.bind("<Tab>", self.confirm_selection)

        # Listbox eventleri
        self.listbox.bind("<Return>", self.confirm_selection)
        self.listbox.bind("<Double-Button-1>", self.on_select)
        self.listbox.bind("<Tab>", self.confirm_selection)
        self.listbox.bind("<Down>", self.move_down)
        self.listbox.bind("<Up>", self.move_up)

    def show_all(self):
        self.listbox.delete(0, tk.END)
        for v in self.values:
            self.listbox.insert(tk.END, v)
        self.listbox.pack(fill="x")

    def get_values(self):
        return self.values
    def set_values(self, new_values):
        self.values = new_values
    @property
    def master_values(self):
        return self.values
    @master_values.setter
    def master_values(self, new_values):
        self.values = new_values

    def on_keyrelease(self, event):
        text = self.var.get().lower()
        if text == "":
            self.listbox.pack_forget()
            return
        matches = [v for v in self.values if text in v.lower()]
        self.listbox.delete(0, tk.END)
        if matches:
            for m in matches:
                self.listbox.insert(tk.END, m)
            self.listbox.pack(fill="x")
        else:
            self.listbox.pack_forget()

    def fill_linked_fields(self, value):
        if not (self.kart_table and self.linked_fields and value):
            return
        for child in self.kart_table.get_children():
            tur, ad, b, f, k = self.kart_table.item(child, "values")
            if value == f"{tur} ({ad})":
                self.linked_fields["birim"].delete(0, "end")
                self.linked_fields["birim"].insert(0, b)
                self.linked_fields["fiyat"].delete(0, "end")
                self.linked_fields["fiyat"].insert(0, f)
                self.linked_fields["kdv"].delete(0, "end")
                self.linked_fields["kdv"].insert(0, k)
                break

    def on_select(self, event):
        if not self.listbox.curselection():
            return
        index = self.listbox.curselection()[0]
        value = self.listbox.get(index)
        self.var.set(value)
        self.listbox.pack_forget()
        self.fill_linked_fields(value)
        if self.next_widget:
            self.next_widget.focus_set()

    def confirm_selection(self, event):
        if self.listbox.curselection():
            index = self.listbox.curselection()[0]
            value = self.listbox.get(index)
        else:
            value = self.var.get()
        self.var.set(value)
        self.listbox.pack_forget()
        self.fill_linked_fields(value)
        if self.next_widget:
            self.next_widget.focus_set()
        return "break"

    def focus_listbox(self, event):
        if self.listbox.size() > 0:
            self.listbox.focus_set()
            self.listbox.selection_clear(0, tk.END)
            self.listbox.selection_set(0)
            self.listbox.activate(0)
        return "break"

    def move_down(self, event):
        idx = self.listbox.curselection()[0] if self.listbox.curselection() else -1
        if idx < self.listbox.size() - 1:
            self.listbox.selection_clear(0, tk.END)
            self.listbox.selection_set(idx + 1)
            self.listbox.activate(idx + 1)
        return "break"

    def move_up(self, event):
        idx = self.listbox.curselection()[0] if self.listbox.curselection() else 0
        if idx > 0:
            self.listbox.selection_clear(0, tk.END)
            self.listbox.selection_set(idx - 1)
            self.listbox.activate(idx - 1)
        return "break"

    def get(self):
        return self.var.get()

    def set(self, value):
        self.var.set(value)
        self.listbox.pack_forget()
        if value:
            self.fill_linked_fields(value)

    def update_values(self, new_values):
        self.values = new_values
# ================== END HELPERS ==================








# ================== START CONFIG ==================
import json
import os

AYARLAR_DOSYA = "ayarlar.json"
KARTLAR_DOSYA = "urun_kartlari.json"
MUSTERI_DOSYA = "musteriler.json"

# Varsayılan ayarlar
default_settings = {
    "pencere_boyut": "1000x750",
    "son_sekme": "Fatura"
}

# --- Ayarlar ---
def load_settings():
    if os.path.exists(AYARLAR_DOSYA):
        try:
            with open(AYARLAR_DOSYA, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return default_settings.copy()
    return default_settings.copy()

def save_settings(settings):
    with open(AYARLAR_DOSYA, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=4)
# --- Ayarlar Sonu ---

# --- Ürün Kartları ---
def save_kartlar(kart_table):
    data = []
    for child in kart_table.get_children():
        values = kart_table.item(child, "values")
        data.append(values)
    with open(KARTLAR_DOSYA, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def load_kartlar(kart_table):
    if not os.path.exists(KARTLAR_DOSYA):
        return
    try:
        with open(KARTLAR_DOSYA, "r", encoding="utf-8") as f:
            data = json.load(f)
        for values in data:
            kart_table.insert("", "end", values=values)
    except:
        pass
# --- Ürün Kartları Sonu ---

# --- Müşteriler ---
def save_musteriler(musteri_table):
    data = [musteri_table.item(c, "values") for c in musteri_table.get_children()]
    with open(MUSTERI_DOSYA, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def load_musteriler(musteri_table):
    if not os.path.exists(MUSTERI_DOSYA):
        return
    try:
        with open(MUSTERI_DOSYA, "r", encoding="utf-8") as f:
            data = json.load(f)
        for values in data:
            musteri_table.insert("", "end", values=values)
    except:
        pass
# --- Müşteriler Sonu ---
# ================== END CONFIG ==================

# ================== START LOG ==================
import os

# GUI tarafındaki log_text değişkeni sonradan atanacak
log_text = None  

def log_yaz(mesaj):
    """
    Hem GUI log ekranına hem de konsola yazdırır.
    GUI tarafında log_text varsa oraya da ekler.
    """
    global log_text
    try:
        if log_text is not None:
            log_text.configure(state="normal")
            log_text.insert("end", mesaj + os.linesep)
            log_text.configure(state="disabled")
            log_text.see("end")
    except Exception:
        pass  # GUI yoksa hata verme

    print(mesaj)  # Konsola yazdır
# ================== END LOG ==================



# ================== START QUEUE VIEW ==================
def init_queue_view(frame_parent):
    global queue_table
    from tkinter import ttk

    # Kuyruk tablosu
    queue_frame = tk.LabelFrame(frame_parent, text="Fatura Kuyruğu", padx=10, pady=10)
    queue_frame.pack(fill="both", expand=True, pady=10)

    columns = ("Unvan", "Vergi No", "Açıklama")
    queue_table = ttk.Treeview(queue_frame, columns=columns, show="headings", height=5)

    for col in columns:
        queue_table.heading(col, text=col)
        queue_table.column(col, width=200, anchor="center")

    queue_table.pack(fill="both", expand=True)

def refresh_queue_view():
    """GUI'deki kuyruk tablosunu günceller"""
    global queue_table, fatura_queue
    if not queue_table:
        return

    queue_table.delete(*queue_table.get_children())
    for _, _, _, bilgiler in fatura_queue:
        queue_table.insert("", "end", values=(
            bilgiler.get("unvan", ""),
            bilgiler.get("vergi_no", ""),
            (bilgiler.get("aciklama", "")[:30] + "...") if bilgiler.get("aciklama") else ""
        ))
# ================== END QUEUE VIEW ==================




# ================== START CONTROLLER ==================
def add_kart(tur_combo, ad_entry, birim_entry, fiyat_entry, kdv_combo,
             kart_table, urun_combo, editing_id=None):
    urun_tur = tur_combo.get().strip()
    urun_ad = ad_entry.get().strip()
    birim = birim_entry.get().strip()
    fiyat = fiyat_entry.get().strip()
    kdv = kdv_combo.get().strip()

    if not (urun_tur and urun_ad):
        return

    values = (urun_tur, urun_ad, birim, fiyat, kdv)

    if editing_id:  # Güncelleme modu
        kart_table.item(editing_id, values=values)
    else:  # Yeni ekleme
        kart_table.insert("", "end", values=values)
        full_name = f"{urun_tur} ({urun_ad})"
        current = urun_combo.get_values()
        if full_name not in current:
            urun_combo.set_values(current + [full_name])

    # Alanları sıfırla
    tur_combo.set("")
    ad_entry.delete(0, "end")
    birim_entry.delete(0, "end"); birim_entry.insert(0, "ADET")
    fiyat_entry.delete(0, "end")
    kdv_combo.set("20")

    return None


def add_urun(urun_combo, miktar_entry, birim_entry, fiyat_entry, kdv_entry,
             iskonto_entry, aciklama_entry, urun_table, kart_table, editing_id=None):
    urun = urun_combo.get().strip()
    miktar = miktar_entry.get().strip()
    birim = birim_entry.get().strip()
    fiyat = fiyat_entry.get().strip()
    kdv = kdv_entry.get().strip()
    iskonto = iskonto_entry.get().strip()
    aciklama = aciklama_entry.get().strip()

    if not urun:
        return

    values = (urun, miktar, birim, fiyat, kdv, iskonto, aciklama)

    if editing_id:  # Güncelleme modu
        urun_table.item(editing_id, values=values)
    else:  # Yeni ekleme
        urun_table.insert("", "end", values=values)

    # Alanları sıfırla
    urun_combo.set("")
    miktar_entry.delete(0, "end"); miktar_entry.insert(0, "1")
    birim_entry.delete(0, "end")
    fiyat_entry.delete(0, "end")
    kdv_entry.delete(0, "end"); kdv_entry.insert(0, "20")
    iskonto_entry.delete(0, "end"); iskonto_entry.insert(0, "0")
    aciklama_entry.delete(0, "end")

    return None
# ================== END CONTROLLER ==================

# ================== START EXCEL ==================
import os
from openpyxl import load_workbook

def _to_float_safe(val):
    if val is None or str(val).strip() == "":
        return 0.0
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except:
        return 0.0

def create_temp_excel_from_table(
    urun_table,
    template_file="zirve_excel_şablon.xlsx",
    output_file="test_fatura_zirve.xlsx"
):
    """
    GUI'deki ürün tablosunu alır, Zirve şablonuna göre geçici Excel dosyası oluşturur.
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_dir, template_file)
    output_path = os.path.join(base_dir, output_file)

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Zirve şablon dosyası bulunamadı: {template_path}")

    wb = load_workbook(template_path)
    ws = wb.active

    row_idx = 2
    for item in urun_table.get_children():
        urun, miktar, birim, fiyat, kdv, iskonto, aciklama = urun_table.item(item, "values")

        miktar_f = _to_float_safe(miktar)
        fiyat_f = _to_float_safe(fiyat)
        iskonto_yuzde = _to_float_safe(iskonto)

        # 👇 iskonto tutarını hesapla
        iskonto_tutar = round((fiyat_f * miktar_f) * (iskonto_yuzde / 100.0), 2)

        # 👇 KDV yüzde (boşsa 0)
        try:
            kdv_yuzde = int(float(kdv)) if str(kdv).strip() != "" else 0
        except:
            kdv_yuzde = 0

        # 👇 KDV tutarı hesapla (iskonto sonrası)
        kdv_tutar = round(((fiyat_f * miktar_f) - iskonto_tutar) * (kdv_yuzde / 100.0), 2)

        # 👇 Birim
        birim_yazi = "C62" if miktar_f > 0 else (birim or "")

        # Excel sütunlarına sırayla yaz
        ws[f"A{row_idx}"] = urun
        ws[f"B{row_idx}"] = miktar_f if miktar != "" else None
        ws[f"C{row_idx}"] = birim_yazi
        ws[f"D{row_idx}"] = fiyat_f
        ws[f"E{row_idx}"] = iskonto_yuzde
        ws[f"F{row_idx}"] = iskonto_tutar
        ws[f"G{row_idx}"] = kdv_yuzde
        ws[f"H{row_idx}"] = kdv_tutar
        ws[f"I{row_idx}"] = aciklama

        row_idx += 1

    wb.save(output_path)
    return output_path
# ================== END EXCEL ==================







# ================== START GUI ==================
def gui_main():
    root = tk.Tk()
    root.title("Oto Fatura Programı")
    global urun_table
    global musteri_vkn, musteri_unvan, musteri_adi, musteri_soyadi
    global musteri_vd_sehir, musteri_vd, musteri_adres, musteri_adres_sehir, musteri_ilce
    global fatura_aciklama
    global zirve_sirket_combo, zirve_user, zirve_pass
    global log_text   # 👈 Buraya ekledik
    global headless_var



    ayarlar = load_settings()
    root.geometry(ayarlar.get("pencere_boyut", "1000x750"))

    # --- Sağ Tık Silme Menüsü ---
    def attach_context_delete(table: ttk.Treeview):
        menu = tk.Menu(root, tearoff=0)
        def delete_selected():
            for item in table.selection():
                table.delete(item)
        menu.add_command(label="Sil", command=delete_selected)
        def on_right_click(event):
            iid = table.identify_row(event.y)
            if iid:
                if iid not in table.selection():
                    table.selection_set(iid)
                menu.post(event.x_root, event.y_root)
        table.bind("<Button-3>", on_right_click)
    # --- Sağ Tık Silme Menüsü Sonu ---

    # --- Notebook (Sekmeler) ---
    notebook = ttk.Notebook(root)
    notebook.pack(fill="both", expand=True)
    # --- Notebook Sonu ---

    # ===================================================
    # ==== SEKME 1: FATURA ====
    # ===================================================
    frame_fatura = tk.Frame(notebook, padx=10, pady=10)
    notebook.add(frame_fatura, text="Fatura")
    # --- Fatura İsimlendirme ---
    frame_fatura_isim = tk.LabelFrame(frame_fatura, text="Fatura İsimlendirme", padx=10, pady=10)
    frame_fatura_isim.pack(fill="x", pady=10)

    # Şube seçimi (müşteri çağırıldığında combobox doldurulacak)
    tk.Label(frame_fatura_isim, text="Şube:").grid(row=0, column=0, sticky="w")
    sube_combo = ttk.Combobox(frame_fatura_isim, values=[], width=27)
    sube_combo.grid(row=0, column=1, padx=5, pady=2)

    # Personel adı
    tk.Label(frame_fatura_isim, text="Personel:").grid(row=1, column=0, sticky="w")
    personel_entry = tk.Entry(frame_fatura_isim, width=30)
    personel_entry.grid(row=1, column=1, padx=5, pady=2)

    # İşlem türü (SRV = servis, STŞ = satış, MGZ = mağaza/ye. parça)
    tk.Label(frame_fatura_isim, text="İşlem Türü:").grid(row=2, column=0, sticky="w")
    islem_combo = ttk.Combobox(frame_fatura_isim, values=["SRV", "STŞ", "YEDEK PARÇA"], width=10)
    islem_combo.set("SRV")
    islem_combo.grid(row=2, column=1, padx=5, pady=2)

    # Fatura adını oluştur ve göster
    def olustur_fatura_adi():
        unvan = musteri_unvan.get().strip() or "UNVAN"
        sube = sube_combo.get().strip() or "SUBE"
        personel = personel_entry.get().strip() or "PERSONEL"
        islem = islem_combo.get().strip() or "ISLEM"

        filename = f"{unvan} - {sube} - {personel} {islem}.pdf"
        tk.messagebox.showinfo("Fatura Adı", f"Oluşturulacak fatura adı:\n\n{filename}")
        return filename

    tk.Button(frame_fatura_isim, text="Fatura Adı Oluştur", command=olustur_fatura_adi).grid(row=3, column=0, columnspan=2, pady=10)
    # --- Fatura İsimlendirme Sonu ---

    # --- Müşteri Bilgileri ---
    frame_musteri = tk.LabelFrame(frame_fatura, text="Müşteri Bilgileri", padx=10, pady=10)
    frame_musteri.pack(fill="x", pady=10)

    # Satır 0: VKN/TCKN, Unvan, Adı, Soyadı, Adres (2 satır kaplar), Şubeler
    tk.Label(frame_musteri, text="VKN / TCKN:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
    musteri_vkn = tk.Entry(frame_musteri, width=18)
    musteri_vkn.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="Unvan:").grid(row=0, column=2, sticky="e", padx=5, pady=5)
    musteri_unvan = tk.Entry(frame_musteri, width=20)
    musteri_unvan.grid(row=0, column=3, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="Adı:").grid(row=0, column=4, sticky="e", padx=5, pady=5)
    musteri_adi = tk.Entry(frame_musteri, width=15)
    musteri_adi.grid(row=0, column=5, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="Soyadı:").grid(row=0, column=6, sticky="e", padx=5, pady=5)
    musteri_soyadi = tk.Entry(frame_musteri, width=15)
    musteri_soyadi.grid(row=0, column=7, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="Adres:").grid(row=0, column=8, sticky="ne", padx=5, pady=5)
    musteri_adres = tk.Text(frame_musteri, width=35, height=4)
    musteri_adres.grid(row=0, column=9, rowspan=2, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="Şubeler:").grid(row=0, column=10, sticky="e", padx=5, pady=5)
    musteri_subeler = tk.Entry(frame_musteri, width=25)
    musteri_subeler.grid(row=0, column=11, padx=5, pady=5, sticky="w")

    # Satır 1: Vergi D. Şehir, Vergi Dairesi, Adres Şehir, İlçe
    tk.Label(frame_musteri, text="Vergi D. Şehir:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
    musteri_vd_sehir = tk.Entry(frame_musteri, width=18)
    musteri_vd_sehir.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="Vergi Dairesi:").grid(row=1, column=2, sticky="e", padx=5, pady=5)
    musteri_vd = tk.Entry(frame_musteri, width=20)
    musteri_vd.grid(row=1, column=3, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="Adres Şehir:").grid(row=1, column=4, sticky="e", padx=5, pady=5)
    musteri_adres_sehir = tk.Entry(frame_musteri, width=15)
    musteri_adres_sehir.grid(row=1, column=5, padx=5, pady=5, sticky="w")

    tk.Label(frame_musteri, text="İlçe:").grid(row=1, column=6, sticky="e", padx=5, pady=5)
    musteri_ilce = tk.Entry(frame_musteri, width=15)
    musteri_ilce.grid(row=1, column=7, padx=5, pady=5, sticky="w")

    # --- Fonksiyonlar ---
    def kaydet_musteri():
        values = (
            musteri_vkn.get(),
            musteri_adi.get(),
            musteri_soyadi.get(),
            musteri_unvan.get(),
            musteri_vd_sehir.get(),
            musteri_vd.get(),
            musteri_adres_sehir.get(),
            musteri_ilce.get(),
            musteri_subeler.get(),
            musteri_adres.get("1.0", "end").strip()
        )

        if not values[0]:
            return

        # VKN kontrolü
        for child in musteri_table.get_children():
            mevcut = musteri_table.item(child, "values")
            if mevcut[0] == values[0]:
                # Aynı VKN bulundu → bilgileri getir
                messagebox.showinfo("Bilgi", "Bu VKN/TCKN zaten kayıtlı, bilgileri dolduruldu.")
                musteri_vkn.delete(0, "end"); musteri_vkn.insert(0, mevcut[0])
                musteri_adi.delete(0, "end"); musteri_adi.insert(0, mevcut[1])
                musteri_soyadi.delete(0, "end"); musteri_soyadi.insert(0, mevcut[2])
                musteri_unvan.delete(0, "end"); musteri_unvan.insert(0, mevcut[3])
                musteri_vd_sehir.delete(0, "end"); musteri_vd_sehir.insert(0, mevcut[4])
                musteri_vd.delete(0, "end"); musteri_vd.insert(0, mevcut[5])
                musteri_adres_sehir.delete(0, "end"); musteri_adres_sehir.insert(0, mevcut[6])
                musteri_ilce.delete(0, "end"); musteri_ilce.insert(0, mevcut[7])
                musteri_subeler.delete(0, "end"); musteri_subeler.insert(0, mevcut[8])
                musteri_adres.delete("1.0", "end"); musteri_adres.insert("1.0", mevcut[9])

                # Şube combobox doldur
                subeler_list = [s.strip() for s in mevcut[8].split(",") if s.strip()]
                sube_combo["values"] = subeler_list
                if subeler_list:
                    sube_combo.set(subeler_list[0])
                return

        # Yeni müşteri ekleniyor
        musteri_table.insert("", "end", values=values)
        messagebox.showinfo("Bilgi", "Yeni müşteri kaydedildi.")

        try:
            save_musteriler(musteri_table)
        except Exception:
            pass

        # Şube combobox doldur
        subeler_list = [s.strip() for s in values[8].split(",") if s.strip()]
        sube_combo["values"] = subeler_list
        if subeler_list:
            sube_combo.set(subeler_list[0])

    def musteri_cagir():
        win = tk.Toplevel(root)
        win.title("Müşteri Seç")
        win.geometry("900x400")

        search_var = tk.StringVar()
        tk.Label(win, text="Ara:").pack(anchor="w", padx=5, pady=5)
        search_entry = tk.Entry(win, textvariable=search_var, width=50)
        search_entry.pack(fill="x", padx=5, pady=5)

        columns = (
            "VKN/TCKN", "Adı", "Soyadı", "Unvan", "Vergi D. Şehir",
            "Vergi Dairesi", "Adres Şehir", "İlçe", "Şubeler", "Adres"
        )
        table = ttk.Treeview(win, columns=columns, show="headings", height=12)
        for col in columns:
            table.heading(col, text=col)
            table.column(col, width=120, anchor="center")
        table.pack(fill="both", expand=True, padx=5, pady=5)

        for child in musteri_table.get_children():
            table.insert("", "end", values=musteri_table.item(child, "values"))

        def do_search(*args):
            query = search_var.get().lower()
            table.delete(*table.get_children())
            for child in musteri_table.get_children():
                values = musteri_table.item(child, "values")
                if any(query in str(v).lower() for v in values):
                    table.insert("", "end", values=values)
        search_var.trace("w", do_search)

        def select_customer():
            selected = table.selection()
            if not selected:
                messagebox.showinfo("Uyarı", "Lütfen bir müşteri seçin!")
                return
            values = table.item(selected[0], "values")
            if not values:
                return

            musteri_vkn.delete(0, "end"); musteri_vkn.insert(0, values[0])
            musteri_adi.delete(0, "end"); musteri_adi.insert(0, values[1])
            musteri_soyadi.delete(0, "end"); musteri_soyadi.insert(0, values[2])
            musteri_unvan.delete(0, "end"); musteri_unvan.insert(0, values[3])
            musteri_vd_sehir.delete(0, "end"); musteri_vd_sehir.insert(0, values[4])
            musteri_vd.delete(0, "end"); musteri_vd.insert(0, values[5])
            musteri_adres_sehir.delete(0, "end"); musteri_adres_sehir.insert(0, values[6])
            musteri_ilce.delete(0, "end"); musteri_ilce.insert(0, values[7])
            musteri_subeler.delete(0, "end"); musteri_subeler.insert(0, values[8])
            musteri_adres.delete("1.0", "end"); musteri_adres.insert("1.0", values[9])

            # Şube combobox doldur
            subeler_list = [s.strip() for s in values[8].split(",") if s.strip()]
            sube_combo["values"] = subeler_list
            if subeler_list:
                sube_combo.set(subeler_list[0])

            win.destroy()

        tk.Button(win, text="Seç", command=select_customer).pack(pady=10)

    def temizle_musteri():
        musteri_vkn.delete(0, "end")
        musteri_unvan.delete(0, "end")
        musteri_adi.delete(0, "end")
        musteri_soyadi.delete(0, "end")
        musteri_vd_sehir.delete(0, "end")
        musteri_vd.delete(0, "end")
        musteri_adres_sehir.delete(0, "end")
        musteri_ilce.delete(0, "end")
        musteri_subeler.delete(0, "end")
        musteri_adres.delete("1.0", "end")

    # --- Butonlar (yan yana, sağ tarafa hizalı) ---
    btn_frame = tk.Frame(frame_musteri)
    btn_frame.grid(row=1, column=8, columnspan=4, padx=5, pady=5, sticky="e")

    tk.Button(btn_frame, text="Kaydet", command=kaydet_musteri).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Çağır", command=musteri_cagir).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Temizle", command=temizle_musteri).pack(side="left", padx=5)
    # --- Müşteri Bilgileri Sonu ---



    # --- Ürün Tablosu ---
    frame_urun = tk.LabelFrame(frame_fatura, text="Ürünler", padx=10, pady=10)
    frame_urun.pack(fill="both", expand=True, padx=10, pady=5)

    columns = ("Ürün Adı", "Miktar", "Birim", "Birim Fiyat", "KDV %", "İskonto %", "Açıklama")
    urun_table = ttk.Treeview(frame_urun, columns=columns, show="headings", height=8)
    for col in columns:
        urun_table.heading(col, text=col)
        urun_table.column(col, width=100, anchor="center")
    urun_table.pack(fill="both", expand=True, side="left")

    scroll = ttk.Scrollbar(frame_urun, orient="vertical", command=urun_table.yview)
    urun_table.configure(yscroll=scroll.set)
    scroll.pack(side="right", fill="y")
    attach_context_delete(urun_table)

  

    # --- Ürün Tablosu Sonu ---

    # --- Ürün Ekleme Alanı ---
    frame_add = tk.LabelFrame(frame_fatura, text="Ürün Ekleme Alanı", padx=10, pady=10)
    frame_add.pack(fill="x", pady=10)

    tk.Label(frame_add, text="Ürün:").grid(row=0, column=0, sticky="w")
    urun_combo = AutocompleteCombobox(
        frame_add,
        values=[],
        width=25,
        linked_fields={
            "birim": None,
            "fiyat": None,
            "kdv": None
        },
        kart_table=None   # 👈 kart_table sonra bağlanacak
    )
    urun_combo.grid(row=1, column=0, padx=5)

    tk.Label(frame_add, text="Miktar:").grid(row=0, column=1, sticky="w")
    miktar_entry = tk.Entry(frame_add, width=5); miktar_entry.insert(0, "1")
    miktar_entry.grid(row=1, column=1, padx=5)

    tk.Label(frame_add, text="Birim:").grid(row=0, column=2, sticky="w")
    birim_entry = tk.Entry(frame_add, width=8)
    birim_entry.grid(row=1, column=2, padx=5)

    tk.Label(frame_add, text="Fiyat:").grid(row=0, column=3, sticky="w")
    fiyat_entry = tk.Entry(frame_add, width=10)
    fiyat_entry.grid(row=1, column=3, padx=5)

    tk.Label(frame_add, text="KDV %:").grid(row=0, column=4, sticky="w")
    kdv_entry = tk.Entry(frame_add, width=5); kdv_entry.insert(0, "20")
    kdv_entry.grid(row=1, column=4, padx=5)

    tk.Label(frame_add, text="İskonto %:").grid(row=0, column=5, sticky="w")
    iskonto_entry = tk.Entry(frame_add, width=5); iskonto_entry.insert(0, "0")
    iskonto_entry.grid(row=1, column=5, padx=5)

    tk.Label(frame_add, text="Açıklama:").grid(row=0, column=6, sticky="w")
    aciklama_entry = tk.Entry(frame_add, width=20)
    aciklama_entry.grid(row=1, column=6, padx=5)

    # --- bağlantıları tamamla ---
    urun_combo.linked_fields = {
        "birim": birim_entry,
        "fiyat": fiyat_entry,
        "kdv": kdv_entry
    }

    def on_add_button():
        add_urun(
            urun_combo, miktar_entry, birim_entry, fiyat_entry, kdv_entry,
            iskonto_entry, aciklama_entry, urun_table, kart_table
        )

    tk.Button(frame_add, text="Ürün Ekle", command=on_add_button).grid(row=1, column=7, padx=5)
    tk.Button(frame_add, text="Toplu Ürün Girişi", command=lambda: open_bulk_add_window()).grid(row=1, column=8, padx=5)
    # --- Ürün Ekleme Alanı Sonu ---



    # --- Toplu Ürün Girişi ---
    def open_bulk_add_window():
        bulk_win = tk.Toplevel(root)
        bulk_win.title("Toplu Ürün Girişi")
        bulk_win.geometry("600x400")

        search_var = tk.StringVar()
        selected_items = set()

        product_listbox = tk.Listbox(bulk_win, selectmode=tk.MULTIPLE)
        product_listbox.pack(fill="both", expand=True, padx=5, pady=5)

        # Ürünleri kartlardan al
        all_products = []
        kart_map = {}
        for child in kart_table.get_children():
            tur, ad, b, f, k = kart_table.item(child, "values")
            full_name = f"{tur} ({ad})"
            all_products.append(full_name)
            kart_map[full_name] = (tur, ad, b, f, k)

        # Listeyi doldur (seçimleri koruyarak)
        def refresh_list():
            product_listbox.delete(0, tk.END)
            for p in all_products:
                if search_var.get().lower() in p.lower():
                    product_listbox.insert(tk.END, p)
                    if p in selected_items:
                        product_listbox.selection_set(tk.END)

        # Seçim değişince hafızaya al
        def on_select(event=None):
            selected = [product_listbox.get(i) for i in product_listbox.curselection()]
            for p in all_products:
                if search_var.get().lower() in p.lower():
                    if p in selected:
                        selected_items.add(p)
                    elif p in selected_items:
                        selected_items.remove(p)

        product_listbox.bind("<<ListboxSelect>>", on_select)

        # Arama kutusu
        tk.Entry(bulk_win, textvariable=search_var).pack(fill="x", padx=5, pady=5)
        search_var.trace("w", lambda *args: refresh_list())

        refresh_list()

        def add_selected_products():
            if not selected_items:
                return
            for full_name in selected_items:
                tur, ad, b, f, k = kart_map[full_name]
                urun_table.insert("", "end", values=(
                    full_name, "1", b, f, k, "0", ""
                ))
            bulk_win.destroy()

        tk.Button(bulk_win, text="Faturaya İlave Et", command=add_selected_products).pack(pady=10)
    # --- Fatura Taslak Oluştur Aksiyonu ---

 

    # --- Toplu Ürün Girişi Sonu ---

    # ================== START FATURA GENEL AÇIKLAMA ==================
    # --- Fatura Genel Açıklama ---
    frame_fatura_aciklama = tk.LabelFrame(frame_fatura, text="Fatura Açıklaması", padx=10, pady=10)
    frame_fatura_aciklama.pack(fill="x", pady=10)

    fatura_aciklama = tk.Text(frame_fatura_aciklama, width=100, height=3)
    fatura_aciklama.pack(fill="x", padx=5, pady=5)
    # --- Fatura Genel Açıklama Sonu ---
    # ================== END FATURA GENEL AÇIKLAMA ==================

        # ================== START FATURA TASLAK BUTONU ==================
    btn_frame_fatura = tk.Frame(frame_fatura)
    btn_frame_fatura.pack(pady=5)

    btn_fatura_olustur = tk.Button(
        btn_frame_fatura,
        text="Fatura Taslak Oluştur",
        command=fatura_kes_action,  # artık fonksiyon hazır
        bg="green",
        fg="white"
    )
    btn_fatura_olustur.pack(side="left", padx=5, pady=5)
    # ================== END FATURA TASLAK BUTONU ==================



    # ===================================================
    # ==== SEKME 2: ÜRÜN KARTLARI ====
    # ===================================================
    frame_kartlar = tk.Frame(notebook, padx=10, pady=10)
    notebook.add(frame_kartlar, text="Ürün Kartları")

    # --- Ürün Kart Tablosu ---
    tk.Label(frame_kartlar, text="Ara:").pack(anchor="w", padx=5, pady=2)
    search_var_kart = tk.StringVar()
    search_entry_kart = tk.Entry(frame_kartlar, textvariable=search_var_kart, width=50)
    search_entry_kart.pack(fill="x", padx=5, pady=2)

    kart_columns = ("Ürün Türü", "Ürün Adı", "Birim", "Fiyat", "KDV %")
    kart_table = ttk.Treeview(frame_kartlar, columns=kart_columns, show="headings", height=10)

    for col in kart_columns:
        kart_table.heading(col, text=col)
        kart_table.column(col, width=150, anchor="center")

    kart_table.pack(fill="both", expand=True, pady=10)
    attach_context_delete(kart_table)
    load_kartlar(kart_table)

    # 👇 Ürün Ekleme Alanı'ndaki combobox'a ürünleri yükle
    urun_listesi = []
    for child in kart_table.get_children():
        tur, ad, b, f, k = kart_table.item(child, "values")
        urun_listesi.append(f"{tur} ({ad})")
    urun_combo.set_values(urun_listesi)

    # 👇 bağlantıyı burada yapıyoruz
    urun_combo.kart_table = kart_table



    # --- Ürün Kart Tablosu Arama ---
    def filter_kartlar(*args):
        query = search_var_kart.get().lower()
        kart_table.delete(*kart_table.get_children())
        try:
            with open(KARTLAR_DOSYA, "r", encoding="utf-8") as f:
                data = json.load(f)
            for values in data:
                if any(query in str(v).lower() for v in values):
                    kart_table.insert("", "end", values=values)
        except:
            pass
    search_var_kart.trace("w", filter_kartlar)
    # --- Ürün Kart Tablosu Sonu ---


    # --- Ürün Kart Ekleme Alanı ---
    frame_kart_add = tk.LabelFrame(frame_kartlar, text="Kart Ekleme Alanı", padx=10, pady=10)
    frame_kart_add.pack(fill="x", pady=5)

    tk.Label(frame_kart_add, text="Ürün Türü:").grid(row=0, column=0, sticky="w")
    kart_tur_combo = AutocompleteCombobox(frame_kart_add, values=[
        "YEDEK PARÇA",
        "S.R/O SU ARITMA CİHAZI",
        "SEBİL",
        "R/O BETA SYSTEM - ENDÜSTRİYEL SU ARITMA CİHAZI"
    ], width=30)
    kart_tur_combo.grid(row=1, column=0, padx=5)

    tk.Label(frame_kart_add, text="Ürün Adı:").grid(row=0, column=1, sticky="w")
    kart_ad_entry = tk.Entry(frame_kart_add, width=25)
    kart_ad_entry.grid(row=1, column=1, padx=5)

    tk.Label(frame_kart_add, text="Birim:").grid(row=0, column=2, sticky="w")
    kart_birim_entry = tk.Entry(frame_kart_add, width=10); kart_birim_entry.insert(0, "ADET")
    kart_birim_entry.grid(row=1, column=2, padx=5)

    tk.Label(frame_kart_add, text="Fiyat:").grid(row=0, column=3, sticky="w")
    kart_fiyat_entry = tk.Entry(frame_kart_add, width=10)
    kart_fiyat_entry.grid(row=1, column=3, padx=5)

    tk.Label(frame_kart_add, text="KDV %:").grid(row=0, column=4, sticky="w")
    kart_kdv_combo = AutocompleteCombobox(frame_kart_add, values=["0", "1", "8", "10", "18", "20"], width=5)
    kart_kdv_combo.set("20")
    kart_kdv_combo.grid(row=1, column=4, padx=5)

    def on_kart_add_button():
        add_kart(
            kart_tur_combo, kart_ad_entry, kart_birim_entry,
            kart_fiyat_entry, kart_kdv_combo, kart_table, urun_combo
        )

    tk.Button(frame_kart_add, text="Kart Ekle", command=on_kart_add_button).grid(row=1, column=5, padx=10)
    # --- Ürün Kart Ekleme Alanı Sonu ---

    # ===================================================
    # --- Müşteriler Sekmesi ---
    frame_musteriler = tk.Frame(notebook, padx=10, pady=10)
    notebook.add(frame_musteriler, text="Müşteriler")

    # Arama kutusu
    search_var_musteri = tk.StringVar()
    tk.Label(frame_musteriler, text="Ara:").pack(anchor="w", padx=5, pady=2)
    search_entry_musteri = tk.Entry(frame_musteriler, textvariable=search_var_musteri, width=50)
    search_entry_musteri.pack(fill="x", padx=5, pady=2)

    musteri_columns = (
        "VKN/TCKN", "Adı", "Soyadı", "Unvan", "Vergi D. Şehir",
        "Vergi Dairesi", "Adres Şehir", "İlçe", "Şubeler", "Adres"
    )
    musteri_table = ttk.Treeview(frame_musteriler, columns=musteri_columns, show="headings", height=12)

    for col in musteri_columns:
        musteri_table.heading(col, text=col)
        musteri_table.column(col, width=120, anchor="center")

    musteri_table.pack(fill="both", expand=True, pady=10)
    attach_context_delete(musteri_table)
    load_musteriler(musteri_table)

    # --- Müşteri Tablosu Arama ---
    def filter_musteriler(*args):
        query = search_var_musteri.get().lower()
        musteri_table.delete(*musteri_table.get_children())
        try:
            with open(MUSTERI_DOSYA, "r", encoding="utf-8") as f:
                data = json.load(f)
            for values in data:
                if any(query in str(v).lower() for v in values):
                    musteri_table.insert("", "end", values=values)
        except:
            pass
    search_var_musteri.trace("w", filter_musteriler)

    # --- Müşteri Şubelerini Göster ---
    def show_subeler(event):
        selected = musteri_table.selection()
        if not selected:
            return
        values = musteri_table.item(selected[0], "values")
        if not values or len(values) < 9:
            return

        subeler_str = values[8]
        subeler_list = [s.strip() for s in subeler_str.split(",") if s.strip()]

        win = tk.Toplevel(root)
        win.title(f"{values[3]} - Şubeler")
        win.geometry("400x300")

        tk.Label(win, text=f"Müşteri: {values[3]}").pack(anchor="w", padx=10, pady=5)

        listbox = tk.Listbox(win, height=10)
        listbox.pack(fill="both", expand=True, padx=10, pady=10)

        for s in subeler_list:
            listbox.insert(tk.END, s)

    musteri_table.bind("<Double-1>", show_subeler)

    # --- Müşteri Düzenleme ---
    def duzenle_musteri():
        selected = musteri_table.selection()
        if not selected:
            tk.messagebox.showwarning("Uyarı", "Lütfen düzenlemek için bir müşteri seçin!")
            return

        values = musteri_table.item(selected[0], "values")
        if not values:
            return

        win = tk.Toplevel(root)
        win.title("Müşteri Düzenle")
        win.geometry("600x500")

        labels = [
            "VKN/TCKN", "Adı", "Soyadı", "Unvan", "Vergi D. Şehir",
            "Vergi Dairesi", "Adres Şehir", "İlçe", "Şubeler", "Adres"
        ]
        entries = []

        for i, label in enumerate(labels):
            tk.Label(win, text=label + ":").grid(row=i, column=0, sticky="e", padx=5, pady=5)
            if label == "Adres":
                entry = tk.Text(win, width=50, height=4)
                entry.insert("1.0", values[i])
                entry.grid(row=i, column=1, padx=5, pady=5)
            else:
                entry = tk.Entry(win, width=50)
                entry.insert(0, values[i])
                entry.grid(row=i, column=1, padx=5, pady=5)
            entries.append(entry)

        def save_changes():
            new_values = []
            for i, label in enumerate(labels):
                if label == "Adres":
                    new_values.append(entries[i].get("1.0", "end").strip())
                else:
                    new_values.append(entries[i].get())
            musteri_table.item(selected[0], values=new_values)
            save_musteriler(musteri_table)
            win.destroy()

        tk.Button(win, text="Kaydet", command=save_changes).grid(row=len(labels), column=1, pady=10)

    # --- Müşteri Silme ---
    def sil_musteri():
        selected = musteri_table.selection()
        if not selected:
            tk.messagebox.showwarning("Uyarı", "Lütfen silmek için bir müşteri seçin!")
            return

        answer = tk.messagebox.askyesno("Onay", "Bu müşteriyi silmek istediğinize emin misiniz?")
        if answer:
            for item in selected:
                musteri_table.delete(item)
            save_musteriler(musteri_table)

    # Düzenle ve Sil butonları
    btn_frame = tk.Frame(frame_musteriler)
    btn_frame.pack(pady=5)

    tk.Button(btn_frame, text="Düzenle", command=duzenle_musteri).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Sil", command=sil_musteri).pack(side="left", padx=5)
    # --- Müşteriler Sekmesi Sonu ---

    # ================== START ZİRVE BİLGİLERİ & LOG ==================
    frame_zirve_log = tk.Frame(notebook, padx=10, pady=10)
    notebook.add(frame_zirve_log, text="Zirve Bilgileri & Log")

    # --- Zirve giriş bilgileri ---
    frame_zirve = tk.LabelFrame(frame_zirve_log, text="Zirve Giriş", padx=10, pady=10)
    frame_zirve.pack(fill="x", pady=10)

    tk.Label(frame_zirve, text="Şirket:").grid(row=0, column=0, sticky="e")
    zirve_sirket_combo = ttk.Combobox(frame_zirve, values=[], width=25)
    zirve_sirket_combo.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(frame_zirve, text="Kullanıcı:").grid(row=0, column=2, sticky="e")
    zirve_user = tk.Entry(frame_zirve, width=25)
    zirve_user.grid(row=0, column=3, padx=5, pady=5)

    tk.Label(frame_zirve, text="Şifre:").grid(row=0, column=4, sticky="e")
    zirve_pass = tk.Entry(frame_zirve, width=25, show="*")
    zirve_pass.grid(row=0, column=5, padx=5, pady=5)

    # --- Kaydet & Sil butonları ---
    def kaydet_zirve():
        sirket = zirve_sirket_combo.get().strip()
        kullanici = zirve_user.get().strip()
        sifre = zirve_pass.get().strip()

        if not sirket:
            messagebox.showerror("Hata", "Lütfen şirket adını girin!")
            return

        bilgiler = {}
        if os.path.exists("zirve_bilgileri.json"):
            try:
                with open("zirve_bilgileri.json", "r", encoding="utf-8") as f:
                    bilgiler = json.load(f)
            except:
                pass

        bilgiler[sirket] = {"kullanici": kullanici, "sifre": sifre}

        with open("zirve_bilgileri.json", "w", encoding="utf-8") as f:
            json.dump(bilgiler, f, ensure_ascii=False, indent=4)

        zirve_sirket_combo["values"] = list(bilgiler.keys())
        messagebox.showinfo("Bilgi", f"{sirket} için bilgiler kaydedildi.")

    def sil_zirve():
        sirket = zirve_sirket_combo.get().strip()
        if not sirket or not os.path.exists("zirve_bilgileri.json"):
            return

        with open("zirve_bilgileri.json", "r", encoding="utf-8") as f:
            bilgiler = json.load(f)

        if sirket in bilgiler:
            del bilgiler[sirket]

        with open("zirve_bilgileri.json", "w", encoding="utf-8") as f:
            json.dump(bilgiler, f, ensure_ascii=False, indent=4)

        zirve_user.delete(0, "end")
        zirve_pass.delete(0, "end")
        zirve_sirket_combo["values"] = list(bilgiler.keys())
        messagebox.showinfo("Bilgi", f"{sirket} için bilgiler silindi.")

    # --- Şirket seçildiğinde bilgileri yükle ---
    def sirket_secildi(event=None):
        secilen = zirve_sirket_combo.get().strip()
        if not secilen or not os.path.exists("zirve_bilgileri.json"):
            return
        try:
            with open("zirve_bilgileri.json", "r", encoding="utf-8") as f:
                bilgiler = json.load(f)
            if secilen in bilgiler and isinstance(bilgiler[secilen], dict):
                zirve_user.delete(0, "end")
                zirve_pass.delete(0, "end")
                zirve_user.insert(0, bilgiler[secilen].get("kullanici", ""))
                zirve_pass.insert(0, bilgiler[secilen].get("sifre", ""))
        except Exception as e:
            print("⚠️ Şirket seçilirken hata:", e)

    zirve_sirket_combo.bind("<<ComboboxSelected>>", sirket_secildi)

    tk.Button(frame_zirve, text="Kaydet", command=kaydet_zirve).grid(row=0, column=6, padx=5, pady=5)
    tk.Button(frame_zirve, text="Sil", command=sil_zirve).grid(row=0, column=7, padx=5, pady=5)

    # Headless mod (Sil butonunun yanında)
    headless_var = tk.BooleanVar(value=False)
    tk.Checkbutton(frame_zirve, text="Headless", variable=headless_var).grid(row=0, column=8, padx=5, pady=5)

    # --- Program açıldığında kayıtlı bilgileri yükle ---
    bilgiler = {}
    if os.path.exists("zirve_bilgileri.json"):
        try:
            with open("zirve_bilgileri.json", "r", encoding="utf-8") as f:
                bilgiler = json.load(f)
        except:
            pass
    else:
        bilgiler = {
            "Şirket A": {"kullanici": "kullanici_adi_A", "sifre": "sifreA123"},
            "Şirket B": {"kullanici": "kullanici_adi_B", "sifre": "sifreB456"}
        }
        with open("zirve_bilgileri.json", "w", encoding="utf-8") as f:
            json.dump(bilgiler, f, ensure_ascii=False, indent=4)

    if bilgiler:
        sirketler = list(bilgiler.keys())
        zirve_sirket_combo["values"] = sirketler
        ilk = sirketler[0]
        zirve_sirket_combo.set(ilk)

        if isinstance(bilgiler[ilk], dict):  # ✅ güvenlik kontrolü eklendi
            zirve_user.insert(0, bilgiler[ilk].get("kullanici", ""))
            zirve_pass.insert(0, bilgiler[ilk].get("sifre", ""))
        else:
            print("⚠️ Beklenmedik JSON formatı:", type(bilgiler[ilk]))
    else:
        zirve_sirket_combo["values"] = ["Şirket A", "Şirket B"]

    # --- Log ekranı ---
    frame_log = tk.LabelFrame(frame_zirve_log, text="İşlem Logu", padx=10, pady=10)
    frame_log.pack(fill="both", expand=True, pady=10)

    log_text = tk.Text(frame_log, state="disabled", height=15)
    log_text.pack(fill="both", expand=True)

    # 👇 Kuyruk görüntüleme tablosunu başlat
    init_queue_view(frame_zirve_log)
    # ================== END ZİRVE BİLGİLERİ & LOG ==================




# ================== START MAIN SCRIPT ==================
driver_global = None  # 👈 Chrome'u global tanımladık
fatura_queue = []     # 👈 Fatura kuyruğu burada tanımlı
is_processing = False # 👈 Şu an işlem var mı?
headless_var = None   # 👈 Headless seçeneği (GUI içinde ayarlanır)

def fatura_kes_action():
    from tkinter import messagebox
    import threading, traceback
    global driver_global, fatura_queue, is_processing

    # --- Önce mevcut GUI bilgilerini kuyruğa ekle ---
    try:
        output_file = create_temp_excel_from_table(urun_table)
    except Exception as e:
        messagebox.showerror("Hata", f"Excel oluşturulamadı!\n{e}")
        return

    sirket = zirve_sirket_combo.get().strip()
    kullanici = zirve_user.get().strip()
    sifre = zirve_pass.get().strip()

    if not (sirket and kullanici and sifre):
        messagebox.showwarning("Uyarı", "Lütfen Zirve bilgilerini doldurun!")
        return

    bilgiler = {
        "vergi_no": musteri_vkn.get().strip(),
        "unvan": musteri_unvan.get().strip() or f"{musteri_adi.get()} {musteri_soyadi.get()}",
        "vergi_sehir": musteri_vd_sehir.get().strip(),
        "vergi_dairesi": musteri_vd.get().strip(),
        "adres": musteri_adres.get("1.0", "end").strip(),
        "adres_sehir": musteri_adres_sehir.get().strip(),
        "adres_ilce": musteri_ilce.get().strip(),
        "aciklama": fatura_aciklama.get("1.0", "end").strip(),
        "excel_path": output_file
    }

    fatura_queue.append((sirket, kullanici, sifre, bilgiler))
    log_yaz(f"📌 Fatura kuyruğa eklendi. Toplam: {len(fatura_queue)}")
    refresh_queue_view()  # 👈 Kuyruğu GUI’de güncelle

    # Eğer şu an işlem yapılmıyorsa kuyruğu başlat
    if not is_processing:
        threading.Thread(target=process_queue).start()


def process_queue():
    global driver_global, fatura_queue, is_processing, headless_var
    import traceback
    is_processing = True

    while fatura_queue:
        sirket, kullanici, sifre, bilgiler = fatura_queue.pop(0)
        try:
            log_yaz("🚀 Yeni fatura işleniyor...")

            # Selenium başlat (her faturada sıfırdan aç)
            log_yaz("🌐 Selenium başlatılıyor...")
            service = Service(ChromeDriverManager().install())
            chrome_options = Options()
            try:
                if headless_var is not None and headless_var.get():
                    chrome_options.add_argument("--headless=new")
                    chrome_options.add_argument("--window-size=1920,1080")
            except Exception:
                pass
            driver = webdriver.Chrome(service=service, options=chrome_options)
            try:
                if not (headless_var is not None and headless_var.get()):
                    driver.maximize_window()
            except Exception:
                pass

            # 1. Portala giriş
            login_portal(driver, kullanici, sifre)
            log_yaz(f"✅ {sirket} için Zirve portalına giriş yapıldı.")

            # 2. Fatura oluşturma akışı
            create_invoice_simple(driver, bilgiler)
            log_yaz("📝 Fatura oluşturma ekranı açıldı.")

            check_customer_and_edit(driver, bilgiler)
            log_yaz("🔍 Müşteri kontrolü tamamlandı.")

            upload_products_from_excel(driver, bilgiler["excel_path"])
            log_yaz("📦 Ürünler portala yüklendi.")

            add_invoice_note(driver, bilgiler.get("aciklama", ""))
            log_yaz("📝 Açıklama eklendi.")

            save_and_close_invoice(driver)
            log_yaz("💾 Fatura taslak olarak kaydedildi.")

        except Exception as e:
            log_yaz(f"❌ Hata: {e}")
            log_yaz(traceback.format_exc())

        refresh_queue_view()  # 👈 Her faturadan sonra tabloyu güncelle

    is_processing = False
    log_yaz("✅ Kuyruk tamamlandı, tüm faturalar işlendi.")
    refresh_queue_view()
# ================== END MAIN SCRIPT ==================








print("✅ GUI dosyası çalışıyor")

gui_main()
tk.mainloop()




